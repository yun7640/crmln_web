# -*- coding: utf-8 -*-
"""CRMLN 측정결과 검토 대시보드 — 로그인 + 업로드 검토 + 관리자 + 사용자 설명서 (Flask)."""
import os, json, io, functools
from flask import (Flask, request, session, redirect, url_for, render_template,
                   send_file, send_from_directory, flash)
import auth
import review_engine
import rounds
import stats

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-only-change-me')
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
BASE = os.path.dirname(os.path.abspath(__file__))


def login_required(view):
    @functools.wraps(view)
    def wrapped(*a, **k):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        return view(*a, **k)
    return wrapped


def admin_required(view):
    @functools.wraps(view)
    def wrapped(*a, **k):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        if not auth.is_admin(session['user']):
            return render_template('message.html', user=session.get('user'), is_admin=False,
                                   title='접근 권한 없음', body='관리자만 접근할 수 있는 메뉴입니다.'), 403
        return view(*a, **k)
    return wrapped


@app.context_processor
def inject_nav():
    u = session.get('user')
    return {'user': u, 'is_admin': bool(u and auth.is_admin(u))}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = (request.form.get('username') or '').strip()
        p = request.form.get('password') or ''
        if auth.verify(u, p):
            session['user'] = u
            return redirect(request.args.get('next') or url_for('index'))
        flash('아이디 또는 비밀번호가 올바르지 않습니다.')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('dashboard.html')


@app.route('/view')
@login_required
def view_dashboard():
    return send_from_directory(os.path.join(BASE, 'private'), 'dashboard.html')


@app.route('/manual')
@login_required
def manual():
    return render_template('manual.html')


@app.route('/review', methods=['POST'])
@login_required
def review():
    f = request.files.get('file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return {'error': '.xlsx 측정 파일을 첨부하세요.'}, 400
    try:
        # 회차 라벨은 파일명·시트명에서 추정한다(출력 시트명·제목 표기용, T7).
        out_bytes, summary = review_engine.process(f.read(), filename=f.filename)
    except ValueError as e:
        return {'error': str(e)}, 400
    except Exception as e:
        return {'error': '처리 중 오류: %s' % e}, 500
    bio = io.BytesIO(out_bytes)
    name = os.path.splitext(f.filename)[0] + '_검토.xlsx'
    resp = send_file(bio, as_attachment=True, download_name=name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['X-Review-Summary'] = json.dumps(summary)
    return resp


# ---------- 회차 누적 대시보드 ----------
@app.route('/rounds')
@login_required
def rounds_home():
    payload = rounds.dashboard_payload()
    return render_template('cumulative.html',
                           payload_json=json.dumps(payload, ensure_ascii=False),
                           round_list=rounds.list_rounds(),
                           persistent=rounds.persistent(),
                           backend=rounds.backend())


@app.route('/rounds/data')
@login_required
def rounds_data():
    p = rounds.dashboard_payload()
    p['is_admin'] = bool(auth.is_admin(session.get('user', '')))
    return app.response_class(json.dumps(p, ensure_ascii=False),
                              mimetype='application/json')


@app.route('/selections')
@login_required
def selections_get():
    """⑦탭 — 회차별 **수기** 제출 선택 기준 기록.

    ★ 기록 전용이다. 서버의 채택 로직(`_dcm_pick`/`combo_pick`)은 이 값을 읽지 않으므로
      여기에 무엇을 적어도 검토 결과가 바뀌지 않는다(§0)."""
    return app.response_class(json.dumps(rounds.selection_payload(), ensure_ascii=False),
                              mimetype='application/json')


@app.route('/selections/save', methods=['POST'])
@login_required
def selections_save():
    """회차·모드별 수기 선택 저장(덮어쓰기). 본문은 JSON."""
    d = request.get_json(silent=True) or {}
    label = (d.get('label') or '').strip()
    mode = (d.get('mode') or '').strip().lower()
    ok, msg = rounds.save_selection(label, mode, d.get('samples') or {},
                                    user=session.get('user', ''), note=d.get('note') or '')
    body = json.dumps({'ok': bool(ok), 'message': msg,
                       'label': rounds.canon_label(label), 'mode': mode}, ensure_ascii=False)
    return app.response_class(body, mimetype='application/json', status=200 if ok else 400)


@app.route('/rounds/stats')
@login_required
def rounds_stats():
    """회차 누적 데이터 기반 추가 통계(정밀도·bias 요약·드리프트·제외 index 분포).

    모니터링·진단용이며 반복측정 채택 로직에는 관여하지 않는다(인수인계 §0)."""
    return app.response_class(json.dumps(stats.payload(), ensure_ascii=False),
                              mimetype='application/json')


def _truthy(v):
    return str(v or '').strip().lower() in ('1', 'true', 'on', 'yes')


@app.route('/rounds/preview', methods=['POST'])
@login_required
def rounds_preview():
    """과거 회차 소급 누적(T1) 1단계 — **저장하지 않고** 미리보기만 반환.

    파일명·시트명에서 회차 라벨 후보를 추정하되 자동 저장하지 않는다. 화면에서 사용자가
    라벨을 확인·확정한 뒤 /rounds/add 로 저장한다(인수인계 §10 T1).
    함께 반환하는 것: 기존 라벨 존재 여부, 시드↔소급 계산 값 대조, 연도 제한 판정."""
    f = request.files.get('file')
    if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return {'error': '.xlsx 측정 파일을 첨부하세요.'}, 400
    data = f.read()
    try:
        summary = review_engine.summarize_round(data)
    except ValueError as e:
        return {'error': str(e)}, 400
    except Exception as e:
        return {'error': '요약 중 오류: %s' % e}, 500
    try:
        import openpyxl, io as _io
        sheetnames = openpyxl.load_workbook(_io.BytesIO(data), read_only=True).sheetnames
    except Exception:
        sheetnames = []
    cands = rounds.infer_labels(f.filename, sheetnames)
    picked = (request.form.get('label') or '').strip() or (cands[0]['label'] if cands else '')
    date_certain = _truthy(request.form.get('date_certain'))
    status = rounds.label_status(picked) if picked else None
    ok_year, year_msg = rounds.year_guard(picked, date_certain=date_certain) if picked else (False, '')
    return app.response_class(json.dumps({
        'saved': False,
        'filename': f.filename,
        'sheetnames': sheetnames,
        'mode': summary.get('mode'),
        'n_samples': summary.get('n_samples'),
        'n_qc': summary.get('n_qc'),
        'n_exceed': summary.get('n_exceed'),
        'candidates': cands,
        'suggested_label': picked,
        'status': status,
        'year_ok': ok_year,
        'year_message': year_msg,
        'min_year': rounds.min_backfill_year(),
        'seed_compare': rounds.seed_compare(picked, summary) if picked else None,
        'summary': summary,
        'note': ('추정 라벨은 참고용입니다. 반드시 사용자가 회차를 확인·확정한 뒤 저장하십시오. '
                 '시드 값과 소급 계산 값이 다르면 덮어쓰지 않고 병기합니다.'),
    }, ensure_ascii=False), mimetype='application/json')


@app.route('/rounds/add', methods=['POST'])
@login_required
def rounds_add():
    """측정 파일 업로드 → 회차 **누적 저장만** 하고 JSON을 반환한다.

    ★ 사용자 확정(2026-07-28): **검토 파일 생성은 왼쪽 '자동 검토' 패널(/review)에서만** 한다.
      ⑥탭 회차 누적분석은 **QC bias 경향성 분석 전용**이다.
      사유 — 이미 CRMLN에 제출한 회차는 선택 결과를 다시 검토할 필요가 없다.
      (종전에는 여기서 검토 엑셀까지 만들어 돌려주었다. 그 경로를 제거했다.)
    reference=1 이면 과거 자료 소급 누적(참고용)으로 저장한다 — 반환 형태는 동일하다."""
    f = request.files.get('file')
    label = (request.form.get('label') or '').strip()
    reference = _truthy(request.form.get('reference'))
    date_certain = _truthy(request.form.get('date_certain'))
    if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return {'error': '.xlsx 측정 파일을 첨부하세요.'}, 400
    if not label:
        return {'error': '회차 라벨(예: 2026.7)을 입력하세요.'}, 400
    if reference and not _truthy(request.form.get('confirm')):
        return {'error': '소급 저장은 회차 라벨 확인 후에만 가능합니다(확인 체크 필요).'}, 400
    data = f.read()
    try:
        summary = review_engine.summarize_round(data)
    except ValueError as e:
        return {'error': str(e)}, 400
    except Exception as e:
        return {'error': '요약 중 오류: %s' % e}, 500
    ok, msg = rounds.add_round(label, summary, user=session.get('user', ''),
                               date=(request.form.get('date') or '').strip(),
                               reference=reference, date_certain=date_certain)
    if not ok:
        return {'error': msg}, 400 if reference else 500
    body = {'stored': True, 'label': rounds.canon_label(label), 'mode': summary['mode'],
            'n_samples': summary.get('n_samples'), 'n_exceed': summary.get('n_exceed'),
            'message': msg, 'review_file': False,
            'note': '누적 저장만 했습니다. 검토 파일이 필요하면 왼쪽 "자동 검토" 패널을 사용하십시오.'}
    if reference:
        cmp_ = rounds.seed_compare(label, summary)
        body.update({'reference': True, 'seed_compare': cmp_,
                     'warn': ('시드 값과 다른 항목이 %d개 있습니다. 덮어쓰지 않고 병기합니다.'
                              % cmp_['n_diff']) if cmp_['n_diff'] else ''})
    resp = app.response_class(json.dumps(body, ensure_ascii=False), mimetype='application/json')
    # HTTP 헤더는 latin-1만 허용 → 한글은 \uXXXX(ASCII)로 인코딩(JS JSON.parse가 복원).
    # 본문이 JSON이 된 뒤에도 기존 화면이 이 헤더를 읽으므로 그대로 둔다.
    resp.headers['X-Round-Result'] = json.dumps(
        {'stored': True, 'label': label, 'mode': summary['mode'],
         'n_samples': summary.get('n_samples'), 'n_exceed': summary.get('n_exceed'),
         'message': msg}, ensure_ascii=True)
    return resp


@app.route('/rounds/delete', methods=['POST'])
@admin_required
def rounds_delete():
    ok, msg = rounds.delete_round(request.form.get('label'), request.form.get('mode') or None)
    if request.form.get('ajax'):
        return {'ok': bool(ok), 'message': msg}, (200 if ok else 500)
    flash(msg)
    return redirect(url_for('rounds_home'))


@app.route('/rounds/export')
@admin_required
def rounds_export():
    bio = io.BytesIO(rounds.export_json().encode('utf-8'))
    return send_file(bio, as_attachment=True, download_name='crmln_rounds_backup.json',
                     mimetype='application/json')


@app.route('/rounds/import', methods=['POST'])
@admin_required
def rounds_import():
    f = request.files.get('file')
    raw = f.read().decode('utf-8') if f else (request.form.get('json') or '')
    ok, msg = rounds.import_json(raw)
    flash(msg)
    return redirect(url_for('rounds_home'))


# ---------- 관리자 ----------
@app.route('/admin')
@admin_required
def admin_home():
    return render_template('admin.html', users=auth.list_users(),
                           persistent=auth.persistent(), export=auth.export_users_json(),
                           backend=auth.backend())


@app.route('/admin/add', methods=['POST'])
@admin_required
def admin_add():
    ok, msg = auth.add_user(request.form.get('username'), request.form.get('password'),
                            admin=bool(request.form.get('admin')))
    flash(msg)
    return redirect(url_for('admin_home'))


@app.route('/admin/reset', methods=['POST'])
@admin_required
def admin_reset():
    ok, msg = auth.set_password(request.form.get('username'), request.form.get('password'))
    flash(msg)
    return redirect(url_for('admin_home'))


@app.route('/admin/role', methods=['POST'])
@admin_required
def admin_role():
    ok, msg = auth.set_admin(request.form.get('username'), bool(request.form.get('admin')))
    flash(msg)
    return redirect(url_for('admin_home'))


@app.route('/admin/delete', methods=['POST'])
@admin_required
def admin_delete():
    u = request.form.get('username')
    if u == session.get('user'):
        flash('현재 로그인한 본인 계정은 삭제할 수 없습니다.')
    else:
        ok, msg = auth.delete_user(u)
        flash(msg)
    return redirect(url_for('admin_home'))


@app.route('/healthz')
def healthz():
    return {'ok': True}


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 8000)), debug=False)
