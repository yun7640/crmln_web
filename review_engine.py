# -*- coding: utf-8 -*-
"""CRMLN 측정 엑셀 업로드 → 검토 파일 생성.
전략: 사용자 검토 템플릿(assets/select_template.xlsx)에 업로드 측정값을 '주입'하여
동적 선택 시트(제출결과_선택검토 → 2026.7_결과선택; 수식·조건부서식·드롭다운 100% 보존)를
그대로 재현하고, 정적 요약 시트(2026.7 측정결과 검토)를 추가한다."""
import io, os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

MEMBER = {'TC': ('±1%', 'pct', 1.0), 'BF': ('±2%', 'pct', 2.0),
          'HDL': ('±1 mg/dL', 'mgdl', 1.0), 'LDL': ('±2%', 'pct', 2.0)}
D1 = (5, 6, 7); D2 = (17, 18, 19)   # R1,R2,R3 열 (Day1 / Day2)
BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE, 'assets', 'select_template.xlsx')
TMPL_MS = '2026-07_결과정리'          # 템플릿 측정 시트명
SEL_SRC = '제출결과_선택검토'          # 템플릿 선택 시트명
SEL_DST = '2026.7_결과선택'           # 출력 선택 시트명
DEFAULT_OPTION = '종합 (BF+HDL 상대편차) · 균형 [기본]'
UC_SHEET = 'HDLC UC 검토'      # (구)2026.7 측정결과 검토 — HDL-C UC / β-정량 검토
DCM_SHEET = 'HDLC DCM 검토'    # HDL-C DCM 검토
DCM_HDL_LIM = 1.0             # CRMLN member: HDL-C DCM bias ±1.0 mg/dL, imprecision SD ≤1.0 mg/dL (PS0126 참조)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _median(a):
    """반복 n개의 median. n=3이면 종전 구현(sorted[1])과 결과가 같고,
    짝수 n(2026.7 DCM의 4반복)에서는 가운데 두 값의 평균을 돌려준다.
    ★ 종전에는 `sorted(a)[1]`로 고정되어 있어 4반복에서 median이 2번째 작은 값이 되었다."""
    s = sorted(a)
    n = len(s)
    if n == 0:
        raise ValueError('빈 반복측정값')
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2.0


def combo_pick(BF, HDL, LDL):
    mB, mH = _median(BF), _median(HDL)
    disc = [abs(BF[k] - mB) / mB * 100 + abs(HDL[k] - mH) / mH * 100 for k in range(3)]
    drop = disc.index(max(disc)); keep = [k for k in range(3) if k != drop]
    m = lambda a: (a[keep[0]] + a[keep[1]]) / 2
    cv = lambda a: (((a[keep[0]] - m(a)) ** 2 + (a[keep[1]] - m(a)) ** 2) / 2) ** 0.5 / m(a) * 100
    return drop, keep, m(BF), m(HDL), m(LDL), cv(LDL)


def _parse(ws):
    """측정 시트에서 CS 검체 행위치와 QC/Control bias(원자료 계산)를 추출."""
    rows = {}; qc = []; an = None; kind = None
    DAY = {1: (4, (5, 6, 7)), 2: (16, (17, 18, 19))}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.strip():
            s = b.strip()
            if s.startswith('BF Control'): an, kind = 'BF', 'ctl'
            elif s.startswith('BF Sample'): an, kind = 'BF', 'smp'
            elif s.startswith('HDL Control'): an, kind = 'HDL', 'ctl'
            elif s.startswith('HDL Sample'): an, kind = 'HDL', 'smp'
            elif s == 'LDL Control': an, kind = 'LDL', 'ctl'
            elif s == 'LDL': an, kind = 'LDL', 'smp'
            elif s == 'QC': an, kind = 'TC', 'qc'
        name = ws.cell(r, 3).value
        if isinstance(name, str): name = name.strip()
        if not name: continue
        if kind in ('qc', 'ctl'):
            analyte = 'TC' if kind == 'qc' else an
            for day, (av_col, rep_cols) in DAY.items():
                A = _num(ws.cell(r, av_col).value)
                reps = [_num(ws.cell(r, c).value) for c in rep_cols]
                if A is None or A == 0 or any(x is None for x in reps):
                    continue
                mean = sum(reps) / 3.0
                qc.append(dict(analyte=analyte, name=name, day=day,
                               biaspct=(mean - A) / A * 100,
                               biasmgdl=(mean - A) if analyte == 'HDL' else None))
        elif kind == 'smp' and name.startswith('CS'):
            rows.setdefault(name, {})[an] = r
    return rows, qc


def _extract_samples(ws, rows):
    out = {}
    for name, amap in rows.items():
        d = {}
        for an in ('BF', 'HDL', 'LDL'):
            rr = amap.get(an)
            if not rr: continue
            per = {}
            for day, cols in ((1, D1), (2, D2)):
                vals = [_num(ws.cell(rr, c).value) for c in cols]
                if all(v is not None for v in vals): per[day] = vals
            d[an] = per
        out[name] = d
    return out


def _verdict(an, biaspct, biasmgdl):
    lbl, mode, lim = MEMBER[an]
    val = biasmgdl if mode == 'mgdl' else biaspct
    if val is None: return ('–', None, lbl)
    ok = abs(val) <= lim + 1e-9
    return ('적합' if ok else '초과', ok, lbl)


def _inject(src_ws, dst_ws, force=False):
    """src → dst 셀값 복사. 병합 부셀·(force=False 시)dst 수식·src 수식은 건너뜀."""
    n = 0
    for row in src_ws.iter_rows():
        for c in row:
            dc = dst_ws.cell(c.row, c.column)
            if isinstance(dc, MergedCell):
                continue
            if not force:
                dv = dc.value
                if isinstance(dv, str) and dv.startswith('='):
                    continue
            v = c.value
            if isinstance(v, str) and v.startswith('='):
                continue
            if v is None:
                continue
            if dc.value != v:
                dc.value = v; n += 1
    return n


def process(in_bytes):
    """입력 xlsx 바이트 → (출력 xlsx 바이트, 요약 dict)."""
    up = openpyxl.load_workbook(io.BytesIO(in_bytes))
    up_do = openpyxl.load_workbook(io.BytesIO(in_bytes), data_only=True)
    ms_name = '결과정리' if '결과정리' in up.sheetnames else (TMPL_MS if TMPL_MS in up.sheetnames else None)
    if ms_name is None:
        raise ValueError("'결과정리' 시트를 찾을 수 없습니다. 표준 CRMLN 측정지 형식이 필요합니다.")
    up_ms = up[ms_name]
    if _is_dcm(up_ms):
        return _process_dcm(up, up_ms)
    rows, qc = _parse(up_ms)
    if not rows:
        raise ValueError("CS 검체(BF/HDL/LDL Sample) 데이터를 찾지 못했습니다. 레이아웃을 확인하세요.")
    samples = _extract_samples(up_ms, rows)

    # 검토 템플릿을 base로 로드
    wb = openpyxl.load_workbook(TEMPLATE)

    # 1) 측정값 주입: 업로드 결과정리 → 템플릿 2026-07_결과정리 (템플릿 수식 보존)
    _inject(up_ms, wb[TMPL_MS], force=False)
    # RESULT 원자료도 이번 회차 값으로 갱신(계산값 스냅샷)
    for sn in ('RESULT(Conc)_DAY1', 'RESULT(Conc)_DAY2'):
        if sn in up_do.sheetnames and sn in wb.sheetnames:
            _inject(up_do[sn], wb[sn], force=True)

    # 2) 선택 시트: 기본 옵션으로 초기화 후 이름 변경
    if SEL_SRC in wb.sheetnames:
        sel = wb[SEL_SRC]
        try:
            sel['C4'] = DEFAULT_OPTION
        except Exception:
            pass
        sel.title = SEL_DST

    # 3) 가이드 시트 텍스트의 옛 시트명 갱신
    if '검토_가이드' in wb.sheetnames:
        g = wb['검토_가이드']
        for row in g.iter_rows():
            for c in row:
                if isinstance(c.value, str) and SEL_SRC in c.value:
                    c.value = c.value.replace(SEL_SRC, SEL_DST)

    # 4) 정적 요약 시트 추가
    _build_review_sheet(wb, samples, qc)

    # 5) Excel 열 때 자동 재계산
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    # 6) 시트 순서
    order = ['검토_가이드', 'RESULT(Conc)_DAY1', 'RESULT(Conc)_DAY2',
             TMPL_MS, SEL_DST, UC_SHEET]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    n_exc = sum(1 for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False)
    summary = {'samples': len(rows), 'qc_rows': len(qc), 'member_exceed': n_exc, 'mode': 'dynamic'}
    return out.read(), summary


def _build_review_sheet(wb, samples, qc):
    if UC_SHEET in wb.sheetnames:
        del wb[UC_SHEET]
    ws = wb.create_sheet(UC_SHEET)
    NAVY, BLUE, GREEN, RED2 = '1F3A5F', '2C6E9B', '1B7F4B', 'C0392B'
    thin = Side(style='thin', color='BBBBBB'); box = Border(thin, thin, thin, thin)

    def C(r, c, v, bold=False, size=10, color='222222', fill=None, align='left', border=False, wrap=False, italic=False):
        x = ws.cell(r, c, v); x.font = Font(name='맑은 고딕', bold=bold, size=size, color=color, italic=italic)
        x.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fill: x.fill = PatternFill('solid', fgColor=fill)
        if border: x.border = box
        return x

    def M(r, c1, c2, v, **k):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2); return C(r, c1, v, **k)

    NC = 8; R = 1
    M(R, 1, NC, 'CRMLN 2026년 7월 HDL-C UC 측정결과 검토', bold=True, size=15, color='FFFFFF', fill=NAVY, align='center'); ws.row_dimensions[R].height = 26; R += 1
    M(R, 1, NC, 'KDCA 진단검사의학 표준검사실(NMRL) · CRMLN member laboratory (Lab 509) · HDL-C / LDL-C(β-quantification)', size=10, color='FFFFFF', fill=BLUE, align='center'); R += 2

    def section(t, sub=''):
        nonlocal R
        M(R, 1, NC, t, bold=True, size=11.5, color='FFFFFF', fill=BLUE); R += 1
        if sub:
            M(R, 1, NC, sub, size=9, color='555555', italic=True, wrap=True); ws.row_dimensions[R].height = 26; R += 1

    # ① QC/Control 정확도
    section('① QC · Control 정확도 판정 (CRMLN member laboratory 기준)',
            '기준: NIST/TC ±1%, BF ±2%, HDL ±1 mg/dL, LDL ±2%. QC·Control은 제출 대상 아님(batch 정확도 확인용).')
    for i, t in enumerate(['항목', '관리물질', 'Day', 'bias(%)', 'bias(mg/dL)', '기준', '판정', '']):
        C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    ws.merge_cells(start_row=R, start_column=7, end_row=R, end_column=8); R += 1
    lbl = {'TC': 'NIST(TC)', 'BF': 'BF Control', 'HDL': 'HDL Control', 'LDL': 'LDL Control'}
    for an in ['TC', 'BF', 'HDL', 'LDL']:
        for q in [x for x in qc if x['analyte'] == an]:
            v, ok, lm = _verdict(an, q['biaspct'], q['biasmgdl'])
            col = GREEN if ok else (RED2 if ok is False else '888888')
            C(R, 1, lbl[an], size=9.5, border=True); C(R, 2, q['name'], size=9.5, border=True)
            C(R, 3, 'Day%d' % q['day'], size=9.5, align='center', border=True)
            C(R, 4, '' if q['biaspct'] is None else round(q['biaspct'], 2), size=9.5, align='center', border=True)
            C(R, 5, '' if q['biasmgdl'] is None else round(q['biasmgdl'], 2), size=9.5, align='center', border=True)
            C(R, 6, lm, size=9.5, align='center', border=True)
            C(R, 7, ('✓ ' if ok else ('⚠ ' if ok is False else '')) + v, bold=True, size=9.5, color=col, align='center', border=True)
            ws.merge_cells(start_row=R, start_column=7, end_row=R, end_column=8); ws.cell(R, 8).border = box; R += 1
    R += 1

    # ② 제출 선택 (종합·균형)
    section('② CRMLN 제출 결과 선택 (CS 검체 R1/R2/R3 → 채택 2반복, 기본: 종합 BF+HDL 균형)',
            '동적 선택·옵션 비교는 [2026.7_결과선택] 시트 참조. QC/Control 미제출 · BF·HDL·LDL 동일 R index 잠금.')
    for i, t in enumerate(['검체', 'Day', '제외R', '채택R', 'BF채택', 'HDL채택', 'LDL채택', 'LDL cv%']):
        C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    R += 1
    for name in sorted(samples):
        d = samples[name]
        for day in (1, 2):
            BF = d.get('BF', {}).get(day); HDL = d.get('HDL', {}).get(day); LDL = d.get('LDL', {}).get(day)
            if not (BF and HDL and LDL): continue
            drop, keep, sB, sH, sL, cvL = combo_pick(BF, HDL, LDL)
            C(R, 1, name, size=9, border=True); C(R, 2, 'Day%d' % day, size=9, align='center', border=True)
            C(R, 3, 'R%d' % (drop + 1), size=9, align='center', color=RED2, border=True)
            C(R, 4, 'R%d·R%d' % (keep[0] + 1, keep[1] + 1), size=9, align='center', color=GREEN, border=True)
            C(R, 5, round(sB, 2), bold=True, size=9, align='center', border=True)
            C(R, 6, round(sH, 2), bold=True, size=9, align='center', border=True)
            C(R, 7, round(sL, 2), bold=True, size=9, align='center', border=True)
            C(R, 8, round(cvL, 2), size=9, align='center', border=True); R += 1
    R += 1

    # ③ 종합 고찰
    exc = [q for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False]
    exc_txt = ', '.join('%s %s' % (q['analyte'], q['name']) for q in exc) or '없음'
    section('③ 종합 고찰')
    for t in [
        '· CS 검체는 정밀도 기반(median 이상치 제외)으로 2반복 채택 → 제출(QC·Control 미제출).',
        '· QC·Control member 기준 초과 항목: %s.' % exc_txt,
        '· 제출값은 CDC 참조법 회신 전 잠정이며, 최종 판정은 검토자 확인 후 확정.',
        '· 선택 기준을 바꿔 비교하려면 [2026.7_결과선택] 시트의 C4 옵션(드롭다운)을 사용하십시오.',
    ]:
        M(R, 1, NC, t, size=10, wrap=True, color='333333'); R += 1
    M(R, 1, NC, '※ 본 시트는 업로드 파일로부터 자동 생성됨. 공식 CRMLN 인증 판정은 CDC 평가보고서(PS)에 따름.', size=8.5, color='888888', italic=True); R += 1

    for i, w in enumerate([13, 13, 7, 12, 12, 10, 10, 10]): ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.sheet_view.showGridLines = False
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ============================================================
#  HDL-C DCM (Designated Comparison Method) — 자동 감지·검토
#  β-정량과 열 위치(R1/R2/R3 = E/F/G, Q/R/S) 동일, 행 구성·단일 항목만 다름.
#  측정지 5–12행: NIST(5)·CFS21-01(6)=TC / HDL CFS21-01(7)·HDL QC2(8)=HDL /
#                 HDL Sample CS01–CS04(9–12).  판정: TC ±1%, HDL ±1 mg/dL.
# ============================================================
DCM_CV_GUARD = 15.0  # 동일 검체 반복 CV(%)가 이보다 크면 데이터 정렬 오류로 보고 해당 Day 제외
DCM_HEADER_ROW = 4   # 'A.value / R1 / R2 / R3 [/ R4 …]' 헤더가 있는 행

def _is_dcm(ws):
    return str(ws.cell(7, 2).value or '').strip().startswith('HDL Control')

def _cvn(reps):
    """반복 n개의 CV(%). 모집단 SD 기준(엑셀 측정지 표기와 동일 관례)."""
    reps = [x for x in reps if x is not None]
    n = len(reps)
    if n < 2:
        return 999
    m = sum(reps) / float(n)
    return (sum((x - m) ** 2 for x in reps) / float(n)) ** 0.5 / m * 100 if m else 999

_cv3 = _cvn  # 하위 호환(구 이름)

def _dcm_pick(reps):
    """R1…Rn(n≥3) 중 **median 편차가 큰 순으로 n−2개를 제외**하고 2개를 채택한다.

    · n=3 → 1개 제외(종전과 동일), n=4 → 2개 제외. 2026.7 측정지부터 CS 검체가 4반복이다.
    · 채택 기준은 **정밀도(median 근접도)뿐**이다. 값의 크기나 bias 방향으로 고르지 않는다(§0).
    · ★ 동점 처리: 같은 값이 중복 측정되면 median 편차가 여러 개 동점이 된다. 이때 단순히
      index 순으로 자르면 **동일한 두 값이 채택되어 CV가 0으로 나오는** 일이 생긴다(정밀도가
      실제보다 좋아 보임 → §0 위반). 그래서 편차가 같으면 **정렬 순위가 중앙에서 먼 쪽을 먼저
      제외**한다. n이 짝수일 때 이 규칙은 "순위 중앙 2개 채택"(양 극단 제외)과 같아지며,
      중복값이 있어도 인위적으로 CV를 낮추지 않는다.
    반환: (drop[list, 오름차순], keep[list], sel[채택 2개 평균], cv[채택 2개의 CV%])"""
    vals = [x for x in reps if x is not None]
    n = len(vals)
    if n < 2:
        raise ValueError('DCM 반복측정값이 2개 미만입니다.')
    if n == 2:
        keep, drop = [0, 1], []
    else:
        m = _median(vals)
        rank = {k: i for i, k in enumerate(sorted(range(n), key=lambda k: (vals[k], k)))}
        center = (n - 1) / 2.0
        # 편차 큰 순 → 동점이면 순위가 중앙에서 먼 순 → 그래도 동점이면 index 순(재현성 고정)
        order = sorted(range(n),
                       key=lambda k: (-abs(vals[k] - m), -abs(rank[k] - center), k))
        drop = sorted(order[:n - 2])
        keep = sorted(order[n - 2:])
    a, b = vals[keep[0]], vals[keep[1]]
    sel = (a + b) / 2.0
    cv = (((a - sel) ** 2 + (b - sel) ** 2) / 2.0) ** 0.5 / sel * 100 if sel else 0
    return drop, keep, sel, cv


def _dcm_day_cols(ws):
    """Day1/Day2 블록의 열 위치를 **헤더에서 자동 탐지**한다.

    과거에는 열을 하드코딩(Day2 = P/Q/R/S)했는데, 2026.7 측정지는 CS 검체 반복이 4개(R1–R4)로
    늘면서 Day2 블록이 한 칸 밀렸다. 그 결과 **Day2 전체가 오류 없이 조용히 누락**되었다.
    → 헤더 행에서 'A.value' 와 'R1','R2',… 를 찾아 블록마다 열을 결정한다.
    반환: {day: {'a': A.value열, 'r': [R열…], 'label': 구분열, 'name': 검체명열}}"""
    hdr = DCM_HEADER_ROW
    maxc = max(ws.max_column or 0, 30)
    blocks = []
    for c in range(1, maxc + 1):
        v = str(ws.cell(hdr, c).value or '').strip().lower().replace(' ', '')
        if v in ('a.value', 'a.값', 'assignedvalue'):
            rcols = []
            k = c + 1
            while k <= maxc:
                t = str(ws.cell(hdr, k).value or '').strip().upper().replace(' ', '')
                if re.fullmatch(r'R\d+', t or ''):
                    rcols.append(k); k += 1
                else:
                    break
            if len(rcols) >= 3:
                blocks.append({'a': c, 'r': rcols, 'label': c - 2, 'name': c - 1})
    if len(blocks) < 2:      # 헤더를 못 찾으면 종전 하드코딩으로 폴백
        return {1: {'a': 4, 'r': [5, 6, 7], 'label': 2, 'name': 3},
                2: {'a': 16, 'r': [17, 18, 19], 'label': 14, 'name': 15}}
    return {1: blocks[0], 2: blocks[1]}


def _dcm_reps(ws, r, cols):
    """해당 행·블록의 반복값. 뒤쪽 빈 칸은 잘라내 실제 반복 개수만 돌려준다."""
    vals = [_num(ws.cell(r, c).value) for c in cols['r']]
    while vals and vals[-1] is None:
        vals.pop()
    return vals


def _parse_dcm(ws):
    """DCM 측정지(5–12행) → (qc, samples). CV>가드 인 Day는 정렬오류로 제외.
    반복 개수는 측정지에 채워진 만큼(3개 또는 4개) 자동 인식한다."""
    DAY = _dcm_day_cols(ws)
    qc = []
    for r, an in [(5, 'TC'), (6, 'TC'), (7, 'HDL'), (8, 'HDL')]:
        if not str(ws.cell(r, 3).value or '').strip():
            continue
        for day, cols in DAY.items():
            # 관리물질 이름은 Day 블록마다 다르다(Day1 NIST1 / Day2 NIST2) → 블록의 이름 열에서 읽는다.
            name = (str(ws.cell(r, cols['name']).value or '').strip()
                    or str(ws.cell(r, 3).value or '').strip())
            A = _num(ws.cell(r, cols['a']).value)
            reps = _dcm_reps(ws, r, cols)
            if A is None or A == 0 or len(reps) < 3 or any(x is None for x in reps):
                continue
            if _cvn(reps) > DCM_CV_GUARD:
                continue
            mean = sum(reps) / float(len(reps))
            qc.append(dict(analyte=an, name=name, day=day, n_reps=len(reps),
                           biaspct=(mean - A) / A * 100,
                           biasmgdl=(mean - A) if an == 'HDL' else None))
    samples = {}
    for r in range(9, 13):
        name = str(ws.cell(r, 3).value or '').strip()
        if not name.upper().replace(' ', '').startswith('CS'):
            continue
        per = {}
        for day, cols in DAY.items():
            reps = _dcm_reps(ws, r, cols)
            if len(reps) >= 3 and all(x is not None for x in reps) and _cvn(reps) <= DCM_CV_GUARD:
                per[day] = reps
        if per:
            samples[name] = per
    return qc, samples


def _extract_dcm_rows(ws):
    """DCM 측정지(5–12행) → Day별 원자료 행(Excel-mirror 표시용).
    반환: {day: [ {label, name, analyte, is_sample, A, reps[n], mean_n, cv_n, n_reps} ]}"""
    DAY = _dcm_day_cols(ws)
    SPEC = [(5, 'TC', False), (6, 'TC', False), (7, 'HDL', False), (8, 'HDL', False),
            (9, 'HDL', True), (10, 'HDL', True), (11, 'HDL', True), (12, 'HDL', True)]
    out = {1: [], 2: []}
    for r, an, is_s in SPEC:
        if not (str(ws.cell(r, 2).value or '').strip() or str(ws.cell(r, 3).value or '').strip()):
            continue
        for day, cols in DAY.items():
            # 구분·검체명도 Day 블록별 열에서 읽는다(Day2는 NIST2 등 이름이 다름).
            label = (str(ws.cell(r, cols['label']).value or '').strip()
                     or str(ws.cell(r, 2).value or '').strip())
            name = (str(ws.cell(r, cols['name']).value or '').strip()
                    or str(ws.cell(r, 3).value or '').strip())
            if is_s and not name.upper().replace(' ', '').startswith('CS'):
                continue
            reps = _dcm_reps(ws, r, cols)
            if len(reps) < 3 or any(x is None for x in reps):
                continue
            if _cvn(reps) > DCM_CV_GUARD:
                continue
            A = _num(ws.cell(r, cols['a']).value)
            mean_n = sum(reps) / float(len(reps))
            disp = (name or label) if is_s else (label or name)
            out[day].append(dict(label=disp, name=name, analyte=an, is_sample=is_s,
                                 A=A, reps=reps, n_reps=len(reps),
                                 mean3=mean_n, cv3=_cvn(reps)))
    return out

def _process_dcm(up, ms):
    qc, samples = _parse_dcm(ms)
    rows = _extract_dcm_rows(ms)
    if not samples:
        raise ValueError('HDL-C DCM 검체(CS01–CS04)를 찾지 못했습니다. DCM 측정지 형식(5–12행)을 확인하세요.')
    for nm in (DCM_SEL_SHEET, '2026.7 측정결과 검토', UC_SHEET, DCM_GUIDE):
        if nm in up.sheetnames:
            del up[nm]
    ms_title = ms.title
    _build_dcm_review(up, qc, samples, rows)
    # UC 검토파일과 동일하게 가이드·동적 선택 시트도 함께 생성한다.
    try:
        _build_dcm_select(up, ms, qc, rows, ms_title=ms_title)
    except Exception:
        pass          # 선택 시트 생성 실패가 검토 파일 전체를 막지 않도록
    _build_dcm_guide(up, ms_title=ms_title)
    # 순서: 측정/RESULT 뒤에 검토 시트
    order = [DCM_GUIDE, 'RESULT(Conc)_DAY1', 'RESULT(Conc)_DAY2', '결과정리', TMPL_MS,
             DCM_SEL_SHEET, DCM_SHEET]
    up._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 500)
    try:
        up.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    out = io.BytesIO(); up.save(out); out.seek(0)
    n_exc = sum(1 for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False)
    days = sorted({d for v in samples.values() for d in v})
    return out.read(), {'mode': 'dcm', 'samples': len(samples), 'qc_rows': len(qc),
                        'member_exceed': n_exc, 'days': days}

DCM_SEL_SHEET = '2026.7_결과선택'      # DCM 출력 선택 시트명(UC와 동일 명명)
DCM_GUIDE = '검토_가이드'
DCM_OPTIONS = [
    ('정밀도 (median 이상치 제외) · 기본', 'SCORE'),
    ('최소분산쌍 (인접 최소간격)', 'PAIR'),
    ('HDL 높은값 우선 (민감도)', 'DIR_HI'),
    ('HDL 낮은값 우선 (민감도)', 'DIR_LO'),
    ('전체 반복 평균 (비교 참고용)', 'ALL'),
    ('수동 지정 (부록 ③ 표)', 'MAN'),
]
DCM_DEFAULT_OPTION = DCM_OPTIONS[0][0]
# (Day1 헬퍼 시작 열, Day2 헬퍼 시작 열, Day당 헬퍼 폭, 옵션 목록 열)
# 헬퍼는 Day당 15열(n, rank×4, idxOfRank×4, drop×2, 기본drop×2, 수동×2)을 쓴다.
# ★ 두 블록이 겹치면 Day2의 n 수식이 Day1의 '기본 drop2'를 덮어써 **조용히 오답**이 된다(실제로 겪음).
#   H1+HW <= H2, H2+HW <= OPT_C 를 반드시 지킬 것(tools/smoke_headers.py에서 검사).
H_DCM_SEL = (44, 60, 15, 90)


def _build_dcm_select(wb, ms, qc, rows, ms_title='결과정리'):
    """HDL-C DCM 동적 선택 시트 — UC의 [2026.7_결과선택]과 동일한 방식.

    · 측정 시트(결과정리)를 참조하는 **수식 기반**이므로 측정값을 바꾸면 즉시 재계산된다.
    · C4 드롭다운으로 선택 옵션을 바꾸면 채택 셀(노란색)·제외 셀(회색)·평균·CV가 재계산된다.
    · ★ 화면·서버(`_dcm_pick`)와 결과가 어긋나면 안 되므로, 서버가 계산한 채택값을 같은 행에
      함께 적어 두고 **'검증' 열에서 자동 대조**한다(불일치 시 빨강 '불일치'). §0 원칙상
      어떤 옵션도 결과를 유리하게 만들기 위해 쓰지 않는다.
    """
    if DCM_SEL_SHEET in wb.sheetnames:
        del wb[DCM_SEL_SHEET]
    ws = wb.create_sheet(DCM_SEL_SHEET)
    L = get_column_letter
    SRC = "'%s'" % ms_title
    DAY = _dcm_day_cols(ms)
    NAVY, BLUE, GREY = '1F3A5F', '2C6E9B', '6B7683'
    YEL, GRY, RED2, GREEN = 'FFF2A8', 'F2F2F2', 'C0392B', '1B7F4B'
    thin = Side(style='thin', color='BBBBBB'); box = Border(thin, thin, thin, thin)

    def C(r, c, v, bold=False, size=10, color='222222', fill=None, align='left',
          border=False, wrap=False, italic=False, fmt=None):
        x = ws.cell(r, c, v)
        x.font = Font(name='맑은 고딕', bold=bold, size=size, color=color, italic=italic)
        x.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fill:
            x.fill = PatternFill('solid', fgColor=fill)
        if border:
            x.border = box
        if fmt:
            x.number_format = fmt
        return x

    def M(r, c1, c2, v, **k):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        return C(r, c1, v, **k)

    # ── 레이아웃 상수 ────────────────────────────────────────────────
    NREP = 4                    # 표에 항상 R1–R4 자리를 둔다(3반복이면 R4는 공란)
    B0, OFF = 2, 20             # Day1 시작 열(B), Day2 오프셋
    HDR = ['구분', '검체', 'A.value', 'R1', 'R2', 'R3', 'R4',
           'mean(채택)', 'stdev(채택)', 'cv%(채택)', 'bias(%)', 'bias(mg/dL)',
           '기본 채택', 'Δ 평균 (옵션−기본)', '기본 CV%', 'Δ CV%p (옵션−기본)',
           '서버 계산값', '검증']
    NCOL = len(HDR)
    SR0, R0 = 5, 15             # 측정지 5행 ↔ 시트 15행
    NROWS = 8                   # 측정지 5–12행
    SAMPLE_SRC = (9, 10, 11, 12)
    H1, H2, HW, OPT_C = H_DCM_SEL

    def cols(day, base):
        return base + (0 if day == 1 else OFF)

    # ── 헤더 블록 ───────────────────────────────────────────────────
    M(1, B0, B0 + NCOL - 1, 'CRMLN 제출 결과 선택 검토 — HDL-C DCM (2026-07 회차)',
      bold=True, size=15, color='FFFFFF', fill=NAVY, align='center')
    ws.row_dimensions[1].height = 26
    M(2, B0, B0 + NCOL - 1,
      'KDCA NMRL(Lab 509) · Designated Comparison Method · [%s] 시트 양식 기준 · 채택 2반복 = 노란색 표시'
      % ms_title, size=10, color='FFFFFF', fill=BLUE, align='center')

    for i, (lab, _code) in enumerate(DCM_OPTIONS):
        C(4 + i, OPT_C, lab, size=9)
    opt_rng = '$%s$4:$%s$%d' % (L(OPT_C), L(OPT_C), 3 + len(DCM_OPTIONS))

    C(4, B0, '선택 옵션', bold=True, size=10, fill='E8EEF5', border=True)
    C(4, B0 + 1, DCM_DEFAULT_OPTION, bold=True, size=10, fill=YEL, border=True)
    dv = DataValidation(type='list', formula1='=%s' % opt_rng, allow_blank=False)
    ws.add_data_validation(dv); dv.add(ws.cell(4, B0 + 1))
    OPT = '$%s$4' % L(B0 + 1)
    CODE = '$%s$8' % L(B0 + 1)

    C(5, B0, '옵션 설명', bold=True, size=9.5, fill='E8EEF5', border=True)
    M(5, B0 + 1, B0 + NCOL - 1,
      ('=IF({c}="ALL","제외 없이 반복 전체 평균을 사용합니다(선택 미적용·비교 참고용).",'
       'IF({c}="MAN","부록 ③ 표에 검체·Day별로 제외할 replicate를 직접 지정합니다.",'
       'IF({c}="PAIR","정렬 후 인접 간격이 가장 작은 2개(가장 근접한 쌍)를 채택합니다.",'
       'IF(LEFT({c},3)="DIR","방향성 what-if: 정밀도가 아닌 값의 크기로 2개를 채택하므로 민감도 확인 전용입니다.",'
       '"median 대비 편차가 큰 replicate를 n−2개 제외하고 2개를 채택합니다(편차 동점이면 정렬 순위가 중앙에서 먼 쪽을 먼저 제외)."'
       '))))').format(c=CODE), size=9, color='555555', wrap=True)
    ws.row_dimensions[5].height = 28

    C(6, B0, '반복 수', bold=True, size=9.5, fill='E8EEF5', border=True)
    M(6, B0 + 1, B0 + NCOL - 1,
      'CS 검체는 측정지에 채워진 반복 수(3 또는 4)를 그대로 사용하며, n−2개를 제외하고 2개를 채택합니다. '
      'QC·Control은 제출 대상이 아니므로 선택하지 않고 전체 반복 평균을 씁니다.', size=9, color='555555', wrap=True)

    C(7, B0, '참고', bold=True, size=9.5, fill='E8EEF5', border=True)
    M(7, B0 + 1, B0 + NCOL - 1,
      '옵션을 바꾸면 아래 표의 채택 셀(노란색)·제외 셀(회색)·평균·CV·요약이 즉시 재계산됩니다. '
      '측정값을 바꿔도 [%s] 시트를 참조하므로 자동 반영됩니다.' % ms_title, size=9, color='555555', wrap=True)

    C(8, B0, '적용 로직', bold=True, size=9.5, fill='E8EEF5', border=True)
    C(8, B0 + 1, '=' + ''.join('IF(%s=%s$%d,"%s",' % (OPT, '$' + L(OPT_C), 4 + i, code)
                               for i, (_l, code) in enumerate(DCM_OPTIONS[1:], start=1))
      + '"SCORE"' + ')' * (len(DCM_OPTIONS) - 1), size=9, color=GREY, align='center')
    M(8, B0 + 2, B0 + NCOL - 1,
      ('=IF({c}="SCORE","median 대비 편차 최대 replicate n−2개 제외 (기본·제출용)",'
       'IF({c}="PAIR","정렬 인접 간격 최소 쌍 채택",'
       'IF({c}="ALL","제외 없음(전체 반복 평균)",'
       'IF({c}="MAN","부록 ③ 직접 지정","값의 크기 순 상위/하위 2개 채택 — 민감도 전용"))))').format(c=CODE),
      size=9, color='555555')

    M(9, B0, B0 + NCOL - 1,
      ('=IF(LEFT({c},3)="DIR","⚠ 방향성 옵션은 결과를 유리하게 만들기 위한 선택이 아니며 민감도 분석 전용입니다. '
       '제출은 기본(정밀도) 옵션을 사용하십시오.","")').format(c=CODE),
      size=9.5, bold=True, color=RED2)

    M(10, B0, B0 + NCOL - 1, '① 검체·Day별 선택 결과 (%s 시트 양식 · 채택 2반복 = 노란색)' % ms_title,
      bold=True, size=11.5, color='FFFFFF', fill=BLUE)

    # ── Day 표 ──────────────────────────────────────────────────────
    for day in (1, 2):
        b = cols(day, B0)
        dc = DAY[day]
        acol, rcols = dc['a'], dc['r']
        M(12, b, b + 2, 'DAY %d' % day, bold=True, size=11, color='FFFFFF', fill=GREY, align='center')
        M(12, b + 3, b + NCOL - 1,
          "=IF({s}!{c}2=\"\",\"\",{s}!{c}2)".format(s=SRC, c=L(dc['a'] + 5)), size=9, color='555555')
        for i, t in enumerate(HDR):
            C(14, b + i, t, bold=True, size=9, color='FFFFFF', fill=NAVY, align='center', border=True, wrap=True)
        ws.row_dimensions[14].height = 30

        hb = H1 if day == 1 else H2      # 헬퍼 시작 열
        for r in range(R0, R0 + NROWS):
            sr = SR0 + (r - R0)
            is_s = sr in SAMPLE_SRC
            rng = '$%s%d:$%s%d' % (L(b + 3), r, L(b + 6), r)
            # 원자료 참조
            C(r, b, "=IF({s}!{c}{sr}=\"\",\"\",{s}!{c}{sr})".format(s=SRC, c=L(dc['label']), sr=sr), size=9, border=True)
            C(r, b + 1, "=IF({s}!{c}{sr}=\"\",\"\",{s}!{c}{sr})".format(s=SRC, c=L(dc['name']), sr=sr),
              size=9, bold=True, color=BLUE if is_s else '222222', border=True)
            C(r, b + 2, "=IF({s}!{c}{sr}=\"\",\"\",{s}!{c}{sr})".format(s=SRC, c=L(acol), sr=sr),
              size=9, align='center', border=True, fmt='0.00')
            for k in range(NREP):
                v = ("=IF({s}!{c}{sr}=\"\",\"\",{s}!{c}{sr})".format(s=SRC, c=L(rcols[k]), sr=sr)
                     if k < len(rcols) else None)
                C(r, b + 3 + k, v, size=9, align='center', border=True, fmt='0.00')

            n = '$%s%d' % (L(hb), r)
            d1 = '$%s%d' % (L(hb + 9), r)
            d2 = '$%s%d' % (L(hb + 10), r)
            keep_avg = ('AVERAGEIFS({rng},{rng},"<>",{rng},"<>")' if not is_s else None)
            if is_s:
                # 채택 2개 평균: 제외 index 2개를 뺀 나머지의 평균
                pick = ('SUMPRODUCT({rng},--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),'
                        '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))'
                        ) .format(rng=rng) % (L(b + 3), r, d1, L(b + 3), r, d2)
                cnt = ('SUMPRODUCT(--({rng}<>""),--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),'
                       '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))'
                       ).format(rng=rng) % (L(b + 3), r, d1, L(b + 3), r, d2)
                C(r, b + 7, '=IF(%s=0,"",%s/%s)' % (n, pick, cnt), size=9.5, bold=True,
                  align='center', border=True, fmt='0.00', color=GREEN)
                C(r, b + 8, ('=IF({m}="","",SQRT(SUMPRODUCT(({rng}-{m})^2,--({rng}<>""),'
                             '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))/%s))'
                             ).format(rng=rng, m='$%s%d' % (L(b + 7), r)) % (L(b + 3), r, d1, L(b + 3), r, d2, cnt),
                  size=9, align='center', border=True, fmt='0.000')
            else:
                C(r, b + 7, '=IF(COUNT({rng})=0,"",AVERAGE({rng}))'.format(rng=rng),
                  size=9, align='center', border=True, fmt='0.00')
                C(r, b + 8, '=IF(COUNT({rng})<2,"",STDEV({rng}))'.format(rng=rng),
                  size=9, align='center', border=True, fmt='0.000')
            mcell = '$%s%d' % (L(b + 7), r)
            scell = '$%s%d' % (L(b + 8), r)
            acell = '$%s%d' % (L(b + 2), r)
            C(r, b + 9, '=IF(OR({m}="",{m}=0),"",{s}/{m}*100)'.format(m=mcell, s=scell),
              size=9, align='center', border=True, fmt='0.000')
            C(r, b + 10, '=IF(OR({m}="",{a}="",{a}=0),"",({m}-{a})/{a}*100)'.format(m=mcell, a=acell),
              size=9, align='center', border=True, fmt='0.00')
            C(r, b + 11, '=IF(OR({m}="",{a}=""),"",{m}-{a})'.format(m=mcell, a=acell),
              size=9, align='center', border=True, fmt='0.00')
            # 기본(정밀도) 기준 대비 비교
            bd1 = '$%s%d' % (L(hb + 11), r)
            bd2 = '$%s%d' % (L(hb + 12), r)
            if is_s:
                bpick = ('SUMPRODUCT({rng},--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),'
                         '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))').format(rng=rng) % (L(b + 3), r, bd1, L(b + 3), r, bd2)
                bcnt = ('SUMPRODUCT(--({rng}<>""),--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),'
                        '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))').format(rng=rng) % (L(b + 3), r, bd1, L(b + 3), r, bd2)
                C(r, b + 12, '=IF(%s=0,"",%s/%s)' % (n, bpick, bcnt), size=9, align='center',
                  border=True, fmt='0.00', color='888888')
                C(r, b + 13, '=IF(OR({m}="",{bm}=""),"",{m}-{bm})'.format(m=mcell, bm='$%s%d' % (L(b + 12), r)),
                  size=9, align='center', border=True, fmt='0.000')
                bsd = ('SQRT(SUMPRODUCT(({rng}-{bm})^2,--({rng}<>""),'
                       '--(COLUMN({rng})-COLUMN($%s%d)+1<>%s),--(COLUMN({rng})-COLUMN($%s%d)+1<>%s))/%s)'
                       ).format(rng=rng, bm='$%s%d' % (L(b + 12), r)) % (L(b + 3), r, bd1, L(b + 3), r, bd2, bcnt)
                C(r, b + 14, '=IF({bm}="","",{sd}/{bm}*100)'.format(bm='$%s%d' % (L(b + 12), r), sd=bsd),
                  size=9, align='center', border=True, fmt='0.000', color='888888')
                C(r, b + 15, '=IF(OR({cv}="",{bcv}=""),"",{cv}-{bcv})'.format(
                    cv='$%s%d' % (L(b + 9), r), bcv='$%s%d' % (L(b + 14), r)),
                  size=9, align='center', border=True, fmt='0.000')
            else:
                for k in range(12, 16):
                    C(r, b + k, '', size=9, align='center', border=True, fill=GRY)
            # 서버 계산값 · 검증
            C(r, b + 16, None, size=9, align='center', border=True, fmt='0.00', color='888888')
            C(r, b + 17,
              ('=IF(OR({sv}="",{m}=""),"",IF({c}<>"SCORE","(옵션 변경됨)",'
               'IF(ABS({m}-{sv})<=0.005,"일치","불일치")))').format(
                  sv='$%s%d' % (L(b + 16), r), m=mcell, c=CODE),
              size=9, align='center', border=True, bold=True)

        _dcm_sel_helpers(ws, b, hb, R0, NROWS, SR0, SAMPLE_SRC, CODE, L)

    # 서버가 계산한 채택값을 '서버 계산값' 열에 적어 자동 대조(검증 열)
    for day in (1, 2):
        b = cols(day, B0)
        smp = [x for x in (rows.get(day) or []) if x['is_sample']]
        for i, rr in enumerate(smp):
            if i >= len(SAMPLE_SRC):
                break
            _d, _k, sel, _cv = _dcm_pick(rr['reps'])
            ws.cell(R0 + (SAMPLE_SRC[i] - SR0), b + 16, round(sel, 4))

    _dcm_sel_summary(ws, B0, OFF, NCOL, R0, SAMPLE_SRC, SR0, L, C, M, NAVY, BLUE, box)
    _dcm_sel_manual(ws, B0, OFF, R0, SAMPLE_SRC, SR0, L, C, M, NAVY, box, H1, H2)
    _dcm_sel_format(ws, B0, OFF, NCOL, R0, NROWS, SR0, H1, H2, HW, OPT_C, L)
    return ws


def _dcm_sel_helpers(ws, b, hb, R0, NROWS, SR0, SAMPLE_SRC, CODE, L):
    """선택 옵션별 제외 replicate index를 계산하는 헬퍼 열(숨김).

    rank_k = (자기보다 작은 값 수) + (앞쪽 반복 중 같은 값 수) → **중복값이 있어도 순위가 유일**하고
    같은 값이면 앞 index가 낮은 순위를 갖는다(서버 `_dcm_pick`의 정렬 규칙과 동일)."""
    for r in range(R0, R0 + NROWS):
        sr = SR0 + (r - R0)
        is_s = sr in SAMPLE_SRC
        rng = '$%s%d:$%s%d' % (L(b + 3), r, L(b + 6), r)
        ws.cell(r, hb, '=COUNT(%s)' % rng)                                    # n
        for k in range(4):
            cell = '%s%d' % (L(b + 3 + k), r)
            upto = '$%s%d:%s%d' % (L(b + 3), r, L(b + 3 + k), r)
            ws.cell(r, hb + 1 + k,
                    '=IF({c}="",0,COUNTIF({rng},"<"&{c})+COUNTIF({up},{c}))'.format(c=cell, rng=rng, up=upto))
        # idx_of_rank(j)
        for j in range(1, 5):
            f = '=IF({r1}={j},1,IF({r2}={j},2,IF({r3}={j},3,IF({r4}={j},4,0))))'.format(
                j=j, **{'r%d' % (i + 1): '$%s%d' % (L(hb + 1 + i), r) for i in range(4)})
            ws.cell(r, hb + 5 + (j - 1), f)
        n = '$%s%d' % (L(hb), r)
        ix = ['$%s%d' % (L(hb + 5 + j), r) for j in range(4)]     # idx of rank1..rank4
        srt = ['SMALL(%s,%d)' % (rng, j) for j in range(1, 5)]
        if not is_s:
            ws.cell(r, hb + 9, 0); ws.cell(r, hb + 10, 0)
            ws.cell(r, hb + 11, 0); ws.cell(r, hb + 12, 0)
            continue
        # 기본(SCORE) 제외 index — n=4는 순위 1·n 제외, n=3은 median에서 먼 쪽 1개 제외
        lo_far = '({m}-{s1})'.format(m='MEDIAN(%s)' % rng, s1=srt[0])
        hi_far = '({s3}-{m})'.format(m='MEDIAN(%s)' % rng, s3=srt[2])
        score1 = ('=IF({n}=4,{i1},IF({n}=3,IF({hi}>{lo},{i3},IF({lo}>{hi},{i1},MIN({i1},{i3}))),0))'
                  ).format(n=n, i1=ix[0], i3=ix[2], hi=hi_far, lo=lo_far)
        score2 = '=IF({n}=4,{i4},0)'.format(n=n, i4=ix[3])
        ws.cell(r, hb + 11, score1)      # 기본 drop1 (옵션과 무관하게 항상 계산 — 비교 기준)
        ws.cell(r, hb + 12, score2)      # 기본 drop2
        b1, b2 = '$%s%d' % (L(hb + 11), r), '$%s%d' % (L(hb + 12), r)
        # PAIR: 정렬 인접 간격이 최소인 쌍을 채택 → 나머지 제외.
        # ★ 간격 동점(중복 측정·부동소수 오차) 시 어느 쌍을 고르냐가 갈리므로 9자리로 반올림해 비교하고,
        #   동점이면 **가장 중앙 쌍**을 택한다(대시보드 JS `dcmPick`과 반드시 동일해야 함).
        g1 = 'ROUND({s2}-{s1},9)'.format(s1=srt[0], s2=srt[1])
        g2 = 'ROUND({s3}-{s2},9)'.format(s2=srt[1], s3=srt[2])
        g3 = 'ROUND({s4}-{s3},9)'.format(s3=srt[2], s4=srt[3])
        # n=4: 중앙쌍(rank2,3) 우선 → 제외 rank1,4 / 하단쌍 → 제외 rank3,4 / 상단쌍 → 제외 rank1,2
        pair1 = ('IF({n}=4,IF({g2}<=MIN({g1},{g3}),{i1},IF({g1}<={g3},{i3},{i1})),'
                 'IF({g1}<={g2},{i3},{i1}))').format(n=n, g1=g1, g2=g2, g3=g3, i1=ix[0], i3=ix[2])
        pair2 = ('IF({n}=4,IF({g2}<=MIN({g1},{g3}),{i4},IF({g1}<={g3},{i4},{i2})),0)'
                 ).format(n=n, g1=g1, g2=g2, g3=g3, i2=ix[1], i4=ix[3])
        # 방향성: 상위 2개 / 하위 2개 채택
        hi1 = 'IF({n}=4,{i1},{i1})'.format(n=n, i1=ix[0])
        hi2 = 'IF({n}=4,{i2},0)'.format(n=n, i2=ix[1])
        lo1 = 'IF({n}=4,{i4},{i3})'.format(n=n, i3=ix[2], i4=ix[3])
        lo2 = 'IF({n}=4,{i3},0)'.format(n=n, i3=ix[2])
        man1 = 'IFERROR(VALUE(MID($%s%d,2,2)),0)' % (L(hb + 13), r)
        man2 = 'IFERROR(VALUE(MID($%s%d,2,2)),0)' % (L(hb + 14), r)
        ws.cell(r, hb + 9,
                '=IF({c}="ALL",0,IF({c}="MAN",{m1},IF({c}="PAIR",{p1},IF({c}="DIR_HI",{h1},'
                'IF({c}="DIR_LO",{l1},{b1})))))'.format(c=CODE, m1=man1, p1=pair1, h1=hi1, l1=lo1, b1=b1))
        ws.cell(r, hb + 10,
                '=IF({c}="ALL",0,IF({c}="MAN",{m2},IF({c}="PAIR",{p2},IF({c}="DIR_HI",{h2},'
                'IF({c}="DIR_LO",{l2},{b2})))))'.format(c=CODE, m2=man2, p2=pair2, h2=hi2, l2=lo2, b2=b2))


def _dcm_sel_summary(ws, B0, OFF, NCOL, R0, SAMPLE_SRC, SR0, L, C, M, NAVY, BLUE, box):
    """② 제출용 요약 — 검체별 Day1·Day2 채택값·Day차·평균."""
    R = R0 + 10
    M(R, B0, B0 + 7, '② 제출용 요약 (채택 2반복 · Day1–Day2 평균)', bold=True, size=11.5,
      color='FFFFFF', fill=BLUE); R += 1
    for i, t in enumerate(['검체', 'HDL Day1', 'HDL Day2', 'Day차(mg/dL)', 'Day차(%)',
                           '제출값(평균)', 'CV Day1(%)', 'CV Day2(%)']):
        C(R, B0 + i, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    R += 1
    for sr in SAMPLE_SRC:
        rr = R0 + (sr - SR0)
        d1 = '$%s%d' % (L(B0 + 7), rr)
        d2 = '$%s%d' % (L(B0 + OFF + 7), rr)
        C(R, B0, '=IF($%s%d="","",$%s%d)' % (L(B0 + 1), rr, L(B0 + 1), rr), size=9, border=True)
        C(R, B0 + 1, '=IF({a}="","",{a})'.format(a=d1), size=9, align='center', border=True, fmt='0.00')
        C(R, B0 + 2, '=IF({b}="","",{b})'.format(b=d2), size=9, align='center', border=True, fmt='0.00')
        C(R, B0 + 3, '=IF(OR({a}="",{b}=""),"",{b}-{a})'.format(a=d1, b=d2), size=9, align='center', border=True, fmt='0.00')
        C(R, B0 + 4, '=IF(OR({a}="",{b}="",{a}=0),"",({b}-{a})/{a}*100)'.format(a=d1, b=d2),
          size=9, align='center', border=True, fmt='0.00')
        C(R, B0 + 5, '=IF(OR({a}="",{b}=""),"",AVERAGE({a},{b}))'.format(a=d1, b=d2),
          size=10, bold=True, align='center', border=True, fmt='0.00', color='1B7F4B')
        C(R, B0 + 6, '=IF($%s%d="","",$%s%d)' % (L(B0 + 9), rr, L(B0 + 9), rr),
          size=9, align='center', border=True, fmt='0.000')
        C(R, B0 + 7, '=IF($%s%d="","",$%s%d)' % (L(B0 + OFF + 9), rr, L(B0 + OFF + 9), rr),
          size=9, align='center', border=True, fmt='0.000')
        R += 1
    M(R, B0, B0 + 7,
      '제출은 CS 검체만 합니다(QC·Control 미제출). 최종 제출·판정은 CDC 참조법 회신 및 검토자 확인 후 확정합니다. '
      '결과를 유리하게 만들기 위한 옵션 선택은 지양합니다.', size=9, color='555555', italic=True, wrap=True)
    ws.row_dimensions[R].height = 24


def _dcm_sel_manual(ws, B0, OFF, R0, SAMPLE_SRC, SR0, L, C, M, NAVY, box, H1, H2):
    """③ 부록 — '수동 지정' 옵션에서 쓰는 제외 replicate 입력표."""
    R = R0 + 10 + 3 + len(SAMPLE_SRC) + 2
    M(R, B0, B0 + 4, "③ 부록 — 수동 지정 (선택 옵션이 '수동 지정'일 때만 적용)",
      bold=True, size=11, color='FFFFFF', fill=NAVY); R += 1
    M(R, B0, B0 + 4, '검체·Day별로 제외할 replicate를 R1–R4 중에서 고릅니다. 2개를 골라야 채택이 2개가 됩니다.',
      size=9, color='555555'); R += 1
    for i, t in enumerate(['검체', 'Day', '제외 ①', '제외 ②', '비고']):
        C(R, B0 + i, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    R += 1
    dv = DataValidation(type='list', formula1='"R1,R2,R3,R4"', allow_blank=True)
    ws.add_data_validation(dv)
    for day in (1, 2):
        hb = H1 if day == 1 else H2
        for sr in SAMPLE_SRC:
            rr = R0 + (sr - SR0)
            C(R, B0, '=IF($%s%d="","",$%s%d)' % (L(B0 + 1), rr, L(B0 + 1), rr), size=9, border=True)
            C(R, B0 + 1, 'Day%d' % day, size=9, align='center', border=True)
            for k in (0, 1):
                cell = C(R, B0 + 2 + k, None, size=9, align='center', border=True, fill='FFFDF0')
                dv.add(cell)
                # 입력 셀을 헬퍼(hb+13, hb+14)로 연결
                ws.cell(rr, hb + 13 + k, '=$%s%d' % (L(B0 + 2 + k), R))
            C(R, B0 + 4,
              '=IF(COUNTA($%s%d:$%s%d)=2,"","※ 2개를 지정해야 채택 2반복이 됩니다")'
              % (L(B0 + 2), R, L(B0 + 3), R), size=8.5, color='C0392B', border=True)
            R += 1


def _dcm_sel_format(ws, B0, OFF, NCOL, R0, NROWS, SR0, H1, H2, HW, OPT_C, L):
    """조건부 서식 — 채택=노란색, 제외=회색 취소선, 검증 불일치=빨강, bias 한계 초과=빨강.

    ★ 판정 단위 주의: TC(NIST·CFS)는 **±1%**, HDL Control은 **±1 mg/dL** 이다.
      두 행에 같은 열·같은 한계를 적용하면 TC가 잘못 빨강 처리된다(측정지 TC bias는 mg/dL로 1을 쉽게 넘김)."""
    YEL, GRY, RED2 = 'FFF2A8', 'F2F2F2', 'C0392B'
    r1, r2 = R0, R0 + NROWS - 1
    tc_rows = (R0 + (5 - SR0), R0 + (6 - SR0))      # 측정지 5·6행 = NIST·CFS (TC)
    hdl_rows = (R0 + (7 - SR0), R0 + (8 - SR0))     # 측정지 7·8행 = HDL Control
    for day, hb in ((1, H1), (2, H2)):
        b = B0 + (0 if day == 1 else OFF)
        rep = '%s%d:%s%d' % (L(b + 3), r1, L(b + 6), r2)
        d1 = '$%s%d' % (L(hb + 9), r1)
        d2 = '$%s%d' % (L(hb + 10), r1)
        base = 'AND({c}<>"",OR({d1}>0,{d2}>0)'.format(c='%s%d' % (L(b + 3), r1), d1=d1, d2=d2)
        drop = base + ',OR(COLUMN()-COLUMN($%s%d)+1=%s,COLUMN()-COLUMN($%s%d)+1=%s))' % (
            L(b + 3), r1, d1, L(b + 3), r1, d2)
        keep = base + ',COLUMN()-COLUMN($%s%d)+1<>%s,COLUMN()-COLUMN($%s%d)+1<>%s)' % (
            L(b + 3), r1, d1, L(b + 3), r1, d2)
        ws.conditional_formatting.add(rep, FormulaRule(
            formula=[drop], fill=PatternFill('solid', fgColor=GRY),
            font=Font(name='맑은 고딕', size=9, color='999999', strike=True), stopIfTrue=True))
        ws.conditional_formatting.add(rep, FormulaRule(
            formula=[keep], fill=PatternFill('solid', fgColor=YEL),
            font=Font(name='맑은 고딕', size=9, bold=True, color='222222')))
        # 검증 열
        ver = '%s%d:%s%d' % (L(b + 17), r1, L(b + 17), r2)
        ws.conditional_formatting.add(ver, FormulaRule(
            formula=['ISNUMBER(SEARCH("불일치",%s%d))' % (L(b + 17), r1)],
            fill=PatternFill('solid', fgColor='F8D7DA'),
            font=Font(name='맑은 고딕', size=9, bold=True, color=RED2)))
        ws.conditional_formatting.add(ver, FormulaRule(
            formula=['ISNUMBER(SEARCH("일치",%s%d))' % (L(b + 17), r1)],
            font=Font(name='맑은 고딕', size=9, color='1B7F4B')))
        # bias 한계 초과 — TC는 bias(%) ±1%, HDL Control은 bias(mg/dL) ±1 mg/dL
        for rr in tc_rows:
            cell = '%s%d' % (L(b + 10), rr)
            ws.conditional_formatting.add('%s:%s' % (cell, cell), FormulaRule(
                formula=['AND(%s<>"",ABS(%s)>%g)' % (cell, cell, MEMBER['TC'][2])],
                fill=PatternFill('solid', fgColor='F8D7DA'),
                font=Font(name='맑은 고딕', size=9, bold=True, color=RED2)))
        for rr in hdl_rows:
            cell = '%s%d' % (L(b + 11), rr)
            ws.conditional_formatting.add('%s:%s' % (cell, cell), FormulaRule(
                formula=['AND(%s<>"",ABS(%s)>%g)' % (cell, cell, DCM_HDL_LIM)],
                fill=PatternFill('solid', fgColor='F8D7DA'),
                font=Font(name='맑은 고딕', size=9, bold=True, color=RED2)))
        # 열 너비
        ws.column_dimensions[L(b)].width = 15
        ws.column_dimensions[L(b + 1)].width = 14
        for k in range(2, NCOL):
            ws.column_dimensions[L(b + k)].width = 10
        ws.column_dimensions[L(b + 13)].width = 12
        ws.column_dimensions[L(b + 15)].width = 12
        # 헬퍼 열 숨김
        for k in range(HW):
            ws.column_dimensions[L(hb + k)].hidden = True
    ws.column_dimensions[L(OPT_C)].hidden = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(R0, B0)


def _build_dcm_guide(wb, ms_title='결과정리'):
    """HDL-C DCM 검토 가이드 시트 — UC 템플릿의 [검토_가이드]와 동일한 역할."""
    if DCM_GUIDE in wb.sheetnames:
        del wb[DCM_GUIDE]
    ws = wb.create_sheet(DCM_GUIDE)
    NAVY, BLUE = '1F3A5F', '2C6E9B'
    thin = Side(style='thin', color='BBBBBB'); box = Border(thin, thin, thin, thin)

    def C(r, c, v, bold=False, size=10, color='222222', fill=None, align='left',
          border=False, wrap=False, italic=False):
        x = ws.cell(r, c, v)
        x.font = Font(name='맑은 고딕', bold=bold, size=size, color=color, italic=italic)
        x.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fill:
            x.fill = PatternFill('solid', fgColor=fill)
        if border:
            x.border = box
        return x

    def M(r, c1, c2, v, **k):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        return C(r, c1, v, **k)

    R = 1
    M(R, 1, 5, 'CRMLN 반기(1년 2회) 측정결과 검토 템플릿 — HDL-C DCM 사용 가이드',
      bold=True, size=15, color='FFFFFF', fill=NAVY, align='center'); ws.row_dimensions[R].height = 28; R += 1
    M(R, 1, 5, 'KDCA NMRL(Lab 509) · HDL-C Designated Comparison Method(DCM) · 자동 생성',
      size=10, color='FFFFFF', fill=BLUE, align='center'); R += 2

    def sec(t):
        nonlocal R
        M(R, 1, 5, t, bold=True, size=11.5, color='FFFFFF', fill=BLUE); R += 1

    def line(t, color='333333', size=10):
        nonlocal R
        M(R, 1, 5, t, size=size, color=color, wrap=True); ws.row_dimensions[R].height = 20; R += 1

    sec('① 사용 방법 (매 회차)')
    for t in ['1. [%s] 시트의 R 열에 새 회차 측정값을 붙여넣습니다 (CS 검체 R1–R4, QC·Control R1–R3).' % ms_title,
              '2. CS 검체의 채택 2반복 셀이 [%s] 시트에서 자동으로 노란색, 제외 셀은 회색 취소선으로 표시됩니다.' % DCM_SEL_SHEET,
              '3. HDL Control(±1 mg/dL) bias가 기준을 초과하면 해당 셀이 자동으로 빨간색으로 표시됩니다.',
              '4. 값만 바꾸면 즉시 재계산됩니다(수식·조건부 서식 기반). 선택 기준을 바꿔 비교하려면 [%s] 시트의 선택 옵션(C4)을 사용합니다.' % DCM_SEL_SHEET,
              '5. Day1·Day2 열 위치는 측정지 헤더행(A.value / R1…Rn)에서 자동으로 찾으므로 반복 수가 바뀌어도 됩니다.']:
        line(t)
    R += 1

    sec('② CRMLN member laboratory 판정 기준 (자동 빨강 표시 기준)')
    for i, t in enumerate(['항목', '관리물질', '기준', '비고', '']):
        C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    R += 1
    for a, b, c, d in [('TC (NIST/CFS)', 'NIST1·NIST2·CFS21-01', 'bias ±1%', 'SRM 1951 정확도 앵커'),
                       ('HDL-C Control', 'HDL CFS21-01 · HDL QC2', 'bias ±1 mg/dL', '저농도 → mg/dL로 판정'),
                       ('HDL-C 정밀도', 'CS01–CS04', 'imprecision SD ≤1.0 mg/dL', 'PS0126 기준 참조')]:
        for i, t in enumerate([a, b, c, d, '']):
            C(R, i + 1, t, size=9.5, border=True, align='left' if i < 2 else 'center')
        R += 1
    R += 1

    sec('③ 선택(채택) 로직')
    for t in ['· 각 CS 검체·Day에서 반복 n개(측정지에 채워진 수) 중 n−2개를 제외하고 2개를 채택합니다 (4반복이면 2개 제외).',
              '· 제외 기준: median 대비 편차가 큰 replicate부터 제외 — 기본(정밀도) 옵션이며 [%s] 시트에서 옵션을 바꿔 비교할 수 있습니다.' % DCM_SEL_SHEET,
              '· ★ 편차 동점 처리: 같은 값이 중복 측정되면 편차가 동점이 됩니다. 이때는 정렬 순위가 중앙에서 먼 쪽을 먼저 제외합니다.',
              '   그렇게 하지 않으면 동일한 두 값이 채택되어 CV가 0이 되어 정밀도가 실제보다 좋아 보입니다.',
              '· DCM은 HDL-C 단일 항목이므로 UC(β-정량)의 BF·HDL·LDL 동일 R index 잠금은 적용되지 않습니다.',
              '· QC·Control은 제출 대상이 아니므로 선택·표시하지 않습니다(전체 반복 평균 사용).',
              '· 방법론 원칙: 정밀도 기반 선택이며, 결과를 유리하게 만들기 위한 선택은 지양합니다.']:
        line(t)
    R += 1

    sec('④ [%s] 시트 — 선택 옵션별 제출결과 검토' % DCM_SEL_SHEET)
    for t in ['· 목적: [%s] 시트와 동일한 DAY1·DAY2 표 양식으로 선택 결과를 재현하고, 옵션을 바꿔가며 채택값·정밀도를 비교합니다.' % ms_title,
              '· 사용법: [C4] 드롭다운에서 옵션을 고르면 ① 표의 채택 셀(노란색)·제외 셀(회색)·평균·CV와 ② 요약이 즉시 재계산됩니다.',
              "· '수동 지정' 선택 시 부록 ③ 표에서 검체·Day별로 제외할 replicate 2개를 직접 고릅니다.",
              '· ★ 검증 열: 서버가 계산한 채택값과 시트 수식 결과를 자동 대조합니다. "불일치"가 뜨면 시트 수식을 신뢰하지 말고 검토자에게 알리십시오.']:
        line(t)
    R += 1
    for i, t in enumerate(['구분', '선택 옵션', '제외/채택 로직', '', '']):
        C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    R += 1
    rows = [('정밀도 기준 (권장)', DCM_OPTIONS[0][0], 'median 대비 편차가 큰 replicate n−2개 제외 (제출용 기본)'),
            ('', DCM_OPTIONS[1][0], '정렬 후 인접 간격이 가장 작은 2개를 채택'),
            ('방향성 (민감도 전용)', DCM_OPTIONS[2][0], '값이 큰 2개를 채택 — 제출용 아님'),
            ('', DCM_OPTIONS[3][0], '값이 작은 2개를 채택 — 제출용 아님'),
            ('비교/수동', DCM_OPTIONS[4][0], '제외 없이 전체 반복 평균 (선택 미적용)'),
            ('', DCM_OPTIONS[5][0], '부록 ③ 표에서 제외 replicate 직접 지정')]
    for a, b, c in rows:
        C(R, 1, a, size=9.5, bold=bool(a), color='1B7F4B' if '권장' in a else '222222', border=True)
        C(R, 2, b, size=9.5, border=True)
        ws.merge_cells(start_row=R, start_column=3, end_row=R, end_column=5)
        C(R, 3, c, size=9.5, border=True)
        R += 1
    R += 1
    M(R, 1, 5, '⚠ 방향성 옵션은 결과를 유리하게 만들기 위한 선택이 아니며 민감도 분석 전용입니다. '
               '최종 제출·판정은 CDC 참조법 회신 및 검토자 확인 후 확정합니다.',
      bold=True, size=10, color='C0392B', wrap=True); ws.row_dimensions[R].height = 22

    for i, w in enumerate([22, 30, 34, 22, 16]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.sheet_view.showGridLines = False
    return ws


def _build_dcm_review(wb, qc, samples, rows=None):
    if DCM_SHEET in wb.sheetnames:
        del wb[DCM_SHEET]
    ws = wb.create_sheet(DCM_SHEET)
    NAVY, BLUE, GREEN, RED2 = '1F3A5F', '2C6E9B', '1B7F4B', 'C0392B'
    thin = Side(style='thin', color='BBBBBB'); box = Border(thin, thin, thin, thin)

    def C(r, c, v, bold=False, size=10, color='222222', fill=None, align='left', border=False, wrap=False, italic=False):
        x = ws.cell(r, c, v); x.font = Font(name='맑은 고딕', bold=bold, size=size, color=color, italic=italic)
        x.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fill: x.fill = PatternFill('solid', fgColor=fill)
        if border: x.border = box
        return x

    def M(r, c1, c2, v, **k):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2); return C(r, c1, v, **k)

    STRIKE = Font(name='맑은 고딕', size=9, color='C0392B', strike=True)
    NC = 8; R = 1
    M(R, 1, NC, 'CRMLN 2026년 7월 HDL-C DCM 측정결과 검토', bold=True, size=15, color='FFFFFF', fill=NAVY, align='center'); ws.row_dimensions[R].height = 26; R += 1
    M(R, 1, NC, 'KDCA 진단검사의학 표준검사실(NMRL) · CRMLN member laboratory (Lab 509) · HDL-C Designated Comparison Method(DCM)', size=10, color='FFFFFF', fill=BLUE, align='center'); R += 2

    def section(t, sub=''):
        nonlocal R
        M(R, 1, NC, t, bold=True, size=11.5, color='FFFFFF', fill=BLUE); R += 1
        if sub:
            M(R, 1, NC, sub, size=9, color='555555', italic=True, wrap=True); ws.row_dimensions[R].height = 26; R += 1

    # ① QC·Control 정확도
    section('① QC · Control 정확도 판정 (CRMLN member laboratory 기준)',
            '기준: NIST/CFS(TC) ±1%, HDL Control ±1 mg/dL. QC·Control은 제출 대상 아님(batch 정확도 확인용).')
    for i, t in enumerate(['항목', '관리물질', 'Day', 'bias(%)', 'bias(mg/dL)', '기준', '판정', '']):
        C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
    ws.merge_cells(start_row=R, start_column=7, end_row=R, end_column=8); R += 1
    lbl = {'TC': 'TC(SRM/CFS)', 'HDL': 'HDL Control'}
    for an in ['TC', 'HDL']:
        for q in [x for x in qc if x['analyte'] == an]:
            v, ok, lm = _verdict(an, q['biaspct'], q['biasmgdl'])
            col = GREEN if ok else (RED2 if ok is False else '888888')
            C(R, 1, lbl[an], size=9.5, border=True); C(R, 2, q['name'], size=9.5, border=True)
            C(R, 3, 'Day%d' % q['day'], size=9.5, align='center', border=True)
            C(R, 4, '' if q['biaspct'] is None else round(q['biaspct'], 2), size=9.5, align='center', border=True)
            C(R, 5, '' if q['biasmgdl'] is None else round(q['biasmgdl'], 2), size=9.5, align='center', border=True)
            C(R, 6, lm, size=9.5, align='center', border=True)
            C(R, 7, ('✓ ' if ok else ('⚠ ' if ok is False else '')) + v, bold=True, size=9.5, color=col, align='center', border=True)
            ws.merge_cells(start_row=R, start_column=7, end_row=R, end_column=8); ws.cell(R, 8).border = box; R += 1
    R += 1

    # ② 제출 결과 선택 (측정지 재현 Day1/Day2 — HDL-C UC 검토와 동일 방식)
    rows = rows or {1: [], 2: []}
    days_present = [d for d in (1, 2) if rows.get(d)]
    if not days_present:  # 폴백: rows 미제공 시 samples로 최소 구성
        days_present = sorted({d for v in samples.values() for d in v})
    # 측정지의 반복 개수(3 또는 4)를 그대로 따른다.
    nrep = max([rr.get('n_reps', len(rr['reps'])) for d in days_present for rr in (rows.get(d) or [])]
               + [len(v) for s in samples.values() for v in s.values()] or [3])
    section('② CRMLN 제출 결과 선택 (HDL-C DCM 측정지 재현 · R1–R%d → 채택 2반복)' % nrep,
            '정밀도 기반: median 대비 편차가 큰 replicate %d개 제외(취소선) → 나머지 2개 평균 채택. '
            'QC/Control은 제출 대상 아님(미제출).' % max(0, nrep - 2))
    NR = 2  # 반복값 시작 열 offset(구분·지정값 다음)
    hdr = ['구분 · 검체', '지정값'] + ['R%d' % (k + 1) for k in range(nrep)] + \
          ['mean(%d)' % nrep, 'cv%%(%d)' % nrep, '채택(2/%d)' % nrep]
    NC2 = len(hdr)
    for day in days_present:
        M(R, 1, max(NC, NC2), 'Day %d' % day, bold=True, size=10, color='FFFFFF', fill='6B7683', align='center'); R += 1
        for i, t in enumerate(hdr):
            C(R, i + 1, t, bold=True, size=9.5, color='FFFFFF', fill=NAVY, align='center', border=True)
        R += 1
        drow = rows.get(day) or []
        for rr in drow:
            reps = rr['reps']
            C(R, 1, rr['label'], size=9, border=True)
            C(R, 2, '' if rr['A'] in (None, 0) else round(rr['A'], 2), size=9, align='center', border=True)
            if rr['is_sample']:
                drop, keep, sel, cv = _dcm_pick(reps)
                for k in range(nrep):
                    v = '' if k >= len(reps) else round(reps[k], 2)
                    cell = C(R, NR + 1 + k, v, size=9, align='center', border=True,
                             color=(GREEN if k in keep else RED2))
                    if k in drop:
                        cell.font = STRIKE
                C(R, NR + nrep + 1, round(rr['mean3'], 2), size=9, align='center', color='888888', border=True)
                C(R, NR + nrep + 2, round(rr['cv3'], 2), size=9, align='center', color='888888', border=True)
                C(R, NR + nrep + 3, round(sel, 2), bold=True, size=9.5, color=GREEN, align='center', border=True)
            else:
                for k in range(nrep):
                    C(R, NR + 1 + k, '' if k >= len(reps) else round(reps[k], 2),
                      size=9, align='center', border=True)
                C(R, NR + nrep + 1, round(rr['mean3'], 2), size=9, align='center', border=True)
                C(R, NR + nrep + 2, round(rr['cv3'], 2), size=9, align='center', border=True)
                C(R, NR + nrep + 3, '미제출', size=8.5, italic=True, color='888888', align='center', border=True)
            R += 1
        R += 1

    # ③ 종합 고찰
    exc = [q for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False]
    exc_txt = ', '.join('%s %s(Day%d)' % (q['analyte'], q['name'], q['day']) for q in exc) or '없음'
    section('③ 종합 고찰')
    for t in [
        '· HDL-C DCM 검체는 정밀도 기반(median 이상치 제외)으로 2반복 채택 → 제출(QC·Control 미제출).',
        '· QC·Control member 기준 초과 항목: %s.' % exc_txt,
        '· HDL Control 판정은 ±1 mg/dL(절대값), NIST·CFS(TC)는 ±1%.',
        '· 제출값은 CDC 참조법 회신 전 잠정이며, 최종 판정은 검토자 확인 후 확정.',
        '· 동일 검체 반복 CV가 %g%% 초과인 Day는 데이터 정렬 오류로 보고 검토에서 제외했습니다(측정지 확인 권장).' % DCM_CV_GUARD,
    ]:
        M(R, 1, NC, t, size=10, wrap=True, color='333333'); R += 1
    M(R, 1, NC, '※ 본 시트는 업로드 파일로부터 자동 생성됨. 공식 CRMLN 인증 판정은 CDC 평가보고서(PS)에 따름.', size=8.5, color='888888', italic=True); R += 1

    for i, w in enumerate([20, 9] + [9] * nrep + [9, 8, 11]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.sheet_view.showGridLines = False
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ============================================================
#  회차 누적용 요약 추출 — summarize_round(in_bytes) → compact dict
#  서버에 회차별 저장 후 추가 통계 분석(경향·제출값 추이)에 활용.
# ============================================================
def _agg_bias(qc, an, key):
    vals = [q[key] for q in qc if q['analyte'] == an and q.get(key) is not None]
    return round(sum(vals) / len(vals), 3) if vals else None


def _qc_list(qc):
    out = []
    for q in qc:
        _, ok, lm = _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])
        out.append(dict(analyte=q['analyte'], name=q['name'], day=q['day'],
                        biaspct=None if q['biaspct'] is None else round(q['biaspct'], 3),
                        biasmgdl=None if q['biasmgdl'] is None else round(q['biasmgdl'], 3),
                        ok=ok, limit=lm))
    return out


def _summarize_uc(ws):
    rows, qc = _parse(ws)
    samples = _extract_samples(ws, rows)
    subs = []
    for name in sorted(samples):
        d = samples[name]
        for day in (1, 2):
            BF = d.get('BF', {}).get(day); HDL = d.get('HDL', {}).get(day); LDL = d.get('LDL', {}).get(day)
            if not (BF and HDL and LDL):
                continue
            drop, keep, sB, sH, sL, cvL = combo_pick(BF, HDL, LDL)
            subs.append(dict(name=name, day=day, drop=drop + 1, keep=[k + 1 for k in keep],
                             BF=round(sB, 2), HDL=round(sH, 2), LDL=round(sL, 2), cvL=round(cvL, 3),
                             reps=dict(BF=BF, HDL=HDL, LDL=LDL)))
    n_exc = sum(1 for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False)
    return dict(mode='uc', qc=_qc_list(qc),
                qc_bias={'TC': _agg_bias(qc, 'TC', 'biaspct'), 'BF': _agg_bias(qc, 'BF', 'biaspct'),
                         'HDL_mgdl': _agg_bias(qc, 'HDL', 'biasmgdl'), 'LDL': _agg_bias(qc, 'LDL', 'biaspct')},
                samples=subs, n_qc=len(qc), n_exceed=n_exc,
                n_samples=len({s['name'] for s in subs}))


def _summarize_dcm(ws):
    qc, samples = _parse_dcm(ws)
    subs = []
    for name in sorted(samples):
        for day in (1, 2):
            reps = samples[name].get(day)
            if not reps:
                continue
            drop, keep, sel, cv = _dcm_pick(reps)
            # drop은 반복이 4개일 수 있으므로 **리스트**로 저장한다(구 데이터는 정수 — 소비처에서 양쪽 처리).
            subs.append(dict(name=name, day=day, drop=[k + 1 for k in drop],
                             keep=[k + 1 for k in keep], n_reps=len(reps),
                             HDL=round(sel, 2), cv=round(cv, 3), reps=reps))
    n_exc = sum(1 for q in qc if _verdict(q['analyte'], q['biaspct'], q['biasmgdl'])[1] is False)
    return dict(mode='dcm', qc=_qc_list(qc),
                qc_bias={'TC': _agg_bias(qc, 'TC', 'biaspct'), 'HDL_mgdl': _agg_bias(qc, 'HDL', 'biasmgdl')},
                samples=subs, n_qc=len(qc), n_exceed=n_exc,
                n_samples=len({s['name'] for s in subs}))


def summarize_round(in_bytes):
    """업로드 xlsx → 회차 누적 저장용 compact dict(UC/DCM 자동 감지)."""
    up = openpyxl.load_workbook(io.BytesIO(in_bytes), data_only=False)
    ms = '결과정리' if '결과정리' in up.sheetnames else (TMPL_MS if TMPL_MS in up.sheetnames else None)
    if ms is None:
        raise ValueError("'결과정리' 측정 시트를 찾을 수 없습니다.")
    ws = up[ms]
    return _summarize_dcm(ws) if _is_dcm(ws) else _summarize_uc(ws)
