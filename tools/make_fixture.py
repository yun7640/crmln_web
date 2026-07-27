# -*- coding: utf-8 -*-
"""검증용 합성(synthetic) 측정지 생성기 — UC / DCM.

⚠️ 여기서 만드는 값은 **전부 인위적으로 만든 가짜 숫자**입니다.
   실제 CRMLN 측정결과가 아니며, 판정·제출에 절대 사용하지 마십시오.
   목적은 오직 업로드 경로(파싱→검토파일 생성→회차요약)가 죽지 않는지 확인하는 것입니다.

레이아웃 근거(review_engine.py):
  · 측정 시트명: '결과정리'(세로형 fixture는 '결과 취합'·'Sheet2')
  · B열=구분 라벨, C열=검체명
  · Day1: 지정값 D열(4), R1~R3 = E,F,G(5,6,7)
  · Day2(가로형): 지정값 P열(16), R1~R3 = Q,R,S(17,18,19). 4반복이면 Q열(17)로 밀림
  · Day2(세로형): 열은 Day1과 같고 **행이 아래로** 내려감(헤더 16행 또는 17행)
  · DCM 판별: 7행 2열이 'HDL Control'로 시작
  · DCM 행 구성: 블록 시작행 기준 0=NIST(TC), 1=CFS21-01(TC), 2=HDL CFS(HDL), 3=HDL QC2(HDL), 4~7=CS01~CS04

사용:
  python tools/make_fixture.py            # tools/_fixtures/ 에 2개 생성
  python tools/make_fixture.py --out DIR
"""
import argparse
import os

from openpyxl import Workbook

SHEET = '결과정리'


def _put(ws, row, label, name, a1, reps1, a2, reps2):
    """한 행에 Day1/Day2 지정값·3반복을 채운다."""
    if label is not None:
        ws.cell(row, 2, label)
    ws.cell(row, 3, name)
    if a1 is not None:
        ws.cell(row, 4, a1)
    for i, v in enumerate(reps1):
        ws.cell(row, 5 + i, v)
    if a2 is not None:
        ws.cell(row, 16, a2)
    for i, v in enumerate(reps2):
        ws.cell(row, 17 + i, v)


def build_uc(path):
    """UC(β-정량) 합성 측정지. BF/HDL Sample CS01~CS04 + QC/Control 포함."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(1, 1, '합성 검증용 파일 — 실제 측정결과 아님 (synthetic fixture, NOT real data)')

    # QC (TC) : 지정값 대비 약 +0.4% 편향
    _put(ws, 3, 'QC', 'NIST SRM1951', 200.0, [200.6, 200.9, 201.1], 200.0, [200.4, 200.8, 201.2])
    # BF Control
    _put(ws, 5, 'BF Control', 'CFS21-01 BF', 120.0, [120.9, 121.2, 121.5], 120.0, [120.7, 121.0, 121.4])
    # HDL Control — ★ 7행에 두면 _is_dcm()이 DCM으로 오인하므로 반드시 6행에 둔다.
    _put(ws, 6, 'HDL Control', 'HDL QC2', 50.0, [50.3, 50.5, 50.6], 50.0, [50.2, 50.4, 50.7])

    # BF Sample CS01~CS04 (지정값 없음 — 검체)
    ws.cell(9, 2, 'BF Sample')
    bf = {'CS01': ([95.2, 95.6, 97.9], [95.4, 95.7, 95.9]),
          'CS02': ([88.1, 88.4, 88.6], [88.0, 88.5, 90.2]),
          'CS03': ([76.5, 76.8, 77.0], [76.6, 76.9, 77.1]),
          'CS04': ([102.3, 102.7, 102.9], [102.4, 102.8, 103.0])}
    for i, (nm, (r1, r2)) in enumerate(bf.items()):
        _put(ws, 9 + i, None, nm, None, r1, None, r2)

    # HDL Sample CS01~CS04
    ws.cell(14, 2, 'HDL Sample')
    hdl = {'CS01': ([55.4, 55.7, 55.9], [55.5, 55.8, 56.0]),
           'CS02': ([46.2, 46.5, 46.7], [46.3, 46.6, 47.9]),
           'CS03': ([35.8, 36.0, 36.2], [35.9, 36.1, 36.3]),
           'CS04': ([59.9, 60.2, 60.4], [60.0, 60.3, 60.5])}
    for i, (nm, (r1, r2)) in enumerate(hdl.items()):
        _put(ws, 14 + i, None, nm, None, r1, None, r2)

    # LDL Control (B열이 정확히 'LDL Control')
    _put(ws, 19, 'LDL Control', 'LDL QC', 70.0, [70.4, 70.7, 70.9], 70.0, [70.3, 70.6, 71.0])

    # LDL Sample CS01~CS04 — B열이 정확히 'LDL'.
    # β-정량 정합(§4): LDL-C = BF − HDL 이므로 반복별로 그대로 차감해 만든다.
    ws.cell(21, 2, 'LDL')
    for i, nm in enumerate(bf):
        l1 = [round(b - h, 2) for b, h in zip(bf[nm][0], hdl[nm][0])]
        l2 = [round(b - h, 2) for b, h in zip(bf[nm][1], hdl[nm][1])]
        _put(ws, 21 + i, None, nm, None, l1, None, l2)

    wb.save(path)
    return path


def build_dcm(path):
    """DCM 합성 측정지. 7행 B열이 'HDL Control'로 시작해야 DCM으로 인식된다."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(1, 1, '합성 검증용 파일 — 실제 측정결과 아님 (synthetic fixture, NOT real data)')

    _put(ws, 5, 'QC', 'NIST SRM1951', 200.0, [200.5, 200.8, 201.0], 200.0, [200.4, 200.7, 201.1])
    _put(ws, 6, 'QC', 'CFS21-01', 190.0, [190.4, 190.7, 190.9], 190.0, [190.3, 190.6, 191.0])
    _put(ws, 7, 'HDL Control', 'HDL CFS21-01', 50.0, [50.2, 50.4, 50.5], 50.0, [50.1, 50.3, 50.6])
    _put(ws, 8, 'HDL QC2', 'HDL QC2', 40.0, [40.1, 40.3, 40.4], 40.0, [40.2, 40.3, 40.5])

    # CS01~CS04 (9~12행) — PS0126 참조값 근처의 임의값
    cs = {'CS01': ([55.4, 55.7, 55.9], [55.5, 55.8, 56.0]),
          'CS02': ([46.2, 46.5, 46.7], [46.3, 46.6, 46.8]),
          'CS03': ([35.8, 36.0, 36.2], [35.9, 36.1, 36.3]),
          'CS04': ([59.9, 60.2, 60.4], [60.0, 60.3, 60.5])}
    for i, (nm, (r1, r2)) in enumerate(cs.items()):
        _put(ws, 9 + i, 'HDL Sample', nm, None, r1, None, r2)

    wb.save(path)
    return path


def build_dcm4(path):
    """DCM 합성 측정지 — **CS 검체 4반복(R1–R4) 레이아웃**.

    2026.7 실제 측정지처럼 CS 검체에 R4가 추가되어 Day2 블록이 한 칸 밀린 형태다.
    (과거 하드코딩 파서는 이 배치에서 Day2를 통째로 놓쳤다 → 회귀 방지용 fixture)
    헤더 행(4행)에 'A.value / R1…Rn'을 적어 두어야 열 자동 탐지가 동작한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(1, 1, '합성 검증용 파일 — 실제 측정결과 아님 (synthetic fixture, NOT real data)')

    # 헤더: Day1 = D..H, Day2 = Q..U (R4가 생겨 Day2가 한 칸 밀린 배치)
    for col, txt in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3'), (8, 'R4'),
                     (17, 'A.value'), (18, 'R1'), (19, 'R2'), (20, 'R3'), (21, 'R4')]:
        ws.cell(4, col, txt)

    def put(row, label, name2, name, a1, r1, a2, r2):
        if label is not None:
            ws.cell(row, 2, label); ws.cell(row, 15, label)
        ws.cell(row, 3, name); ws.cell(row, 16, name2 or name)
        if a1 is not None:
            ws.cell(row, 4, a1)
        for i, v in enumerate(r1):
            ws.cell(row, 5 + i, v)
        if a2 is not None:
            ws.cell(row, 17, a2)
        for i, v in enumerate(r2):
            ws.cell(row, 18 + i, v)

    # QC/Control은 3반복 유지(실제 측정지와 동일)
    put(5, 'QC', 'NIST2', 'NIST1', 200.0, [200.5, 200.8, 201.0], 300.0, [300.4, 300.7, 301.1])
    put(6, None, 'CFS21-01', 'CFS21-01', 190.0, [190.4, 190.7, 190.9], 190.0, [190.3, 190.6, 191.0])
    put(7, 'HDL Control', 'HDL CFS21-01', 'HDL CFS21-01', 50.0, [50.2, 50.4, 50.5], 50.0, [50.1, 50.3, 50.6])
    put(8, None, 'HDL QC2', 'HDL QC2', 40.0, [40.1, 40.3, 40.4], 40.0, [40.2, 40.3, 40.5])

    # CS 검체는 4반복.
    # CS04는 **중복값이 순위 중앙이 아닌 위치**에 오도록 만든다.
    #   Day1 [59.9, 60.2, 60.4, 60.4] → 중복(60.4)이 상위 2개
    #   Day2 [60.0, 60.0, 60.3, 60.5] → 중복(60.0)이 하위 2개
    # 단순 index 기준으로 자르면 중복값 2개가 채택되어 CV가 0으로 붕괴한다(정밀도가 실제보다
    # 좋아 보임 → §0 위반). 순위 중앙 2개를 채택하면 CV>0 이 유지되어야 한다.
    cs = {'CS01': ([55.4, 55.7, 55.9, 55.8], [55.5, 55.8, 56.0, 55.9]),
          'CS02': ([46.2, 46.5, 46.7, 46.6], [46.3, 46.6, 46.8, 46.7]),
          'CS03': ([35.8, 36.0, 36.2, 36.1], [35.9, 36.1, 36.3, 36.2]),
          'CS04': ([59.9, 60.2, 60.4, 60.4], [60.0, 60.0, 60.3, 60.5])}
    for i, (nm, (r1, r2)) in enumerate(cs.items()):
        put(9 + i, 'HDL Sample' if i == 0 else None, nm, nm, None, r1, None, r2)

    wb.save(path)
    return path


def build_dcm_stacked(path, sheet='결과 취합', hdr2=16):
    """DCM 합성 측정지 — **Day2가 Day1 아래에 세로로 쌓인 레이아웃**(T5 회귀 fixture).

    2025.7(`Sheet2`, Day2 헤더 17행)·2026.1(`결과 취합`, Day2 헤더 16행) 실제 측정지의 배열이다.
    Day2 열은 Day1과 **같지만**(D/E–H) 시작 행이 다르다. 종전 파서는 행을 5–12로 하드코딩해
    Day2를 Day1과 같은 행에서 읽었고, 그 결과 **Day2가 조용히 Day1 사본이 되거나 비었다**.
    → 헤더 행을 실제로 찾아야만 통과하도록 Day1/Day2 값을 서로 다르게 만든다.

    ★ Day2 헤더 간격(hdr2)이 회차마다 다르다는 점이 이 fixture의 핵심이다.
      2025.7=17행, 2026.1=16행이라 '몇 행 아래'로 고정할 수 없다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(1, 1, '합성 검증용 파일 — 실제 측정결과 아님 (synthetic fixture, NOT real data)')

    def block(hdr, tag, qc, cs):
        ws.cell(hdr - 2, 2, tag)                       # 'DAY1' / 'DAY2' 표식
        ws.cell(hdr - 2, 9, '* 실험일 : 2026-01-0%d' % (6 if tag == 'DAY1' else 8))
        for col, txt in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3'), (8, 'R4')]:
            ws.cell(hdr, col, txt)
        for i, (label, name, a, reps) in enumerate(qc):
            r = hdr + 1 + i
            if label:
                ws.cell(r, 2, label)
            ws.cell(r, 3, name)
            ws.cell(r, 4, a)
            for k, v in enumerate(reps):
                ws.cell(r, 5 + k, v)
        for i, (name, reps) in enumerate(cs):
            r = hdr + 5 + i
            if i == 0:
                ws.cell(r, 2, 'HDL Sample x1.09')
            ws.cell(r, 3, name)
            for k, v in enumerate(reps):
                ws.cell(r, 5 + k, v)

    # Day1 / Day2 값을 다르게 두어야 '세로 탐색 실패'가 검사에 걸린다.
    block(4, 'DAY1',
          [('QC', 'NIST1', 154.6, [155.4, 155.2, 154.8]),
           (None, 'NIST2', 244.8, [245.6, 245.0, 244.7]),
           ('HDL Control x1.09', 'HDL CFS21-01', 49.18, [49.29, 48.81, 48.75]),
           (None, 'HDL CFS21-02', 49.73, [49.65, 49.65, 49.71])],
          [('CS01', [56.68, 56.56, 57.04, 56.92]), ('CS02', [41.42, 41.30, 41.78, 40.94]),
           ('CS03', [59.57, 59.63, 59.99, 60.11]), ('CS04', [35.41, 35.29, 35.17, 35.47])])
    block(hdr2, 'DAY2',
          [('QC', 'NIST1', 154.6, [155.6, 155.4, 155.2]),
           (None, 'NIST2', 244.8, [246.2, 245.7, 245.8]),
           ('HDL Control x1.09', 'HDL CFS21-01', 49.18, [49.58, 49.88, 49.34]),
           (None, 'HDL CFS21-02', 49.73, [50.19, 49.82, 49.70])],
          [('CS01', [57.41, 57.29, 57.23, 57.29]), ('CS02', [41.94, 41.76, 42.36, 41.82]),
           ('CS03', [60.36, 60.84, 60.72, 60.30]), ('CS04', [35.98, 36.04, 35.56, 35.56])])

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=os.path.join(os.path.dirname(os.path.abspath(__file__)), '_fixtures'))
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)
    u = build_uc(os.path.join(a.out, 'fixture_UC_합성.xlsx'))
    d = build_dcm(os.path.join(a.out, 'fixture_DCM_합성.xlsx'))
    d4 = build_dcm4(os.path.join(a.out, 'fixture_DCM4_합성.xlsx'))
    # 세로형은 Day2 헤더 간격이 회차마다 달라 두 변형을 모두 만든다(2026.1=16행, 2025.7=17행).
    sv1 = build_dcm_stacked(os.path.join(a.out, 'fixture_DCM_세로_합성.xlsx'), '결과 취합', 16)
    sv2 = build_dcm_stacked(os.path.join(a.out, 'fixture_DCM_세로_Sheet2_합성.xlsx'), 'Sheet2', 17)
    print(u)
    print(d)
    print(d4)
    print(sv1)
    print(sv2)


if __name__ == '__main__':
    main()
