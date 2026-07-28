# -*- coding: utf-8 -*-
"""스모크 검증 ① — 앱 기능 + HTTP 헤더 latin-1 안전성 (크로스 플랫폼).

배경: HTTP 응답 헤더는 latin-1만 허용한다. 헤더에 한글이 그대로 들어가면
      gunicorn(Railway)에서 500이 나지만 Flask 개발서버/test_client는 관대해서 그냥 통과한다.
      → 그래서 여기서는 **헤더 값을 직접 latin-1로 인코딩해보며 명시적으로 검사**한다.
      이 방식이면 Windows·개발서버에서도 동일한 버그를 잡을 수 있다.

실행:
    python tools/smoke_headers.py
종료코드 0=통과, 1=실패.

주의: 이 스크립트가 쓰는 측정 파일은 tools/make_fixture.py가 만든 **합성 데이터**이며
      실제 CRMLN 측정결과가 아니다. 판정·제출에 사용하지 말 것.
"""
import io
import json
import os
import shutil
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

FAILS = []
CHECKS = [0]


def check(name, cond, detail=''):
    CHECKS[0] += 1
    if cond:
        print('  PASS  %s' % name)
    else:
        print('  FAIL  %s  %s' % (name, detail))
        FAILS.append(name)


def assert_latin1_headers(resp, where):
    """응답의 모든 헤더 이름·값이 latin-1로 인코딩 가능한지 검사(한글 헤더 500 재발 방지)."""
    bad = []
    for k, v in resp.headers.items():
        for part, kind in ((k, 'name'), (str(v), 'value')):
            try:
                part.encode('latin-1')
            except UnicodeEncodeError:
                bad.append('%s(%s)=%r' % (k, kind, part))
    check('%s: 헤더 latin-1 안전' % where, not bad, '위반: %s' % bad)


def main():
    tmp = tempfile.mkdtemp(prefix='crmln_smoke_')
    os.environ.update({
        'SECRET_KEY': 'smoke-test-key',
        'ADMIN_USERS': 'admin',
        'ADMIN_PASSWORD': 'smoke-pw',
        'USERS_FILE': os.path.join(tmp, 'users.json'),
        'ROUNDS_FILE': os.path.join(tmp, 'rounds.json'),
    })
    os.environ.pop('DATABASE_URL', None)  # 파일 폴백 경로로 검증

    import make_fixture  # noqa: E402  (tools/ 안)
    fx = os.path.join(tmp, 'fx')
    os.makedirs(fx, exist_ok=True)
    uc_path = make_fixture.build_uc(os.path.join(fx, 'fixture_UC.xlsx'))
    dcm_path = make_fixture.build_dcm(os.path.join(fx, 'fixture_DCM.xlsx'))

    import app as appmod  # noqa: E402
    appmod.app.config['TESTING'] = True
    c = appmod.app.test_client()

    print('[1] 기본 라우트')
    r = c.get('/healthz')
    check('/healthz 200', r.status_code == 200, r.status_code)
    check('/healthz ok=true', r.get_json(silent=True) == {'ok': True}, r.data[:80])
    r = c.get('/')
    check('비로그인 / → 로그인 리다이렉트', r.status_code in (301, 302), r.status_code)

    print('[2] 로그인')
    r = c.post('/login', data={'username': 'admin', 'password': 'wrong'})
    check('오답 비밀번호는 세션 미생성', b'/login' not in r.headers.get('Location', '').encode() or r.status_code == 200,
          r.status_code)
    r = c.post('/login', data={'username': 'admin', 'password': 'smoke-pw'}, follow_redirects=False)
    check('정답 비밀번호 → 리다이렉트', r.status_code in (301, 302), r.status_code)
    r = c.get('/')
    check('로그인 후 / 200', r.status_code == 200, r.status_code)

    print('[3] /review 단발 업로드 (한글 파일명)')
    with open(uc_path, 'rb') as f:
        r = c.post('/review', data={'file': (f, '2026.7 측정결과_한글이름.xlsx')},
                   content_type='multipart/form-data')
    check('/review 200', r.status_code == 200, r.data[:200])
    assert_latin1_headers(r, '/review')
    check('/review 엑셀 반환', r.data[:2] == b'PK', r.data[:8])
    xrs = r.headers.get('X-Review-Summary')
    check('X-Review-Summary 존재·JSON 파싱', bool(xrs) and isinstance(json.loads(xrs), dict), xrs)

    print('[4] /rounds/add 누적 저장 (한글 파일명 + 한글 메시지 헤더)')
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/add',
                   data={'file': (f, '2026.7 측정결과_한글이름.xlsx'), 'label': '2026-07'},
                   content_type='multipart/form-data')
    check('/rounds/add 200', r.status_code == 200, r.data[:300])
    assert_latin1_headers(r, '/rounds/add')
    xrr = r.headers.get('X-Round-Result')
    check('X-Round-Result 존재', bool(xrr), xrr)
    if xrr:
        d = json.loads(xrr)
        check('X-Round-Result stored=True', d.get('stored') is True, d)
        check('X-Round-Result 라벨 정규화 2026-07→2026.7', d.get('label') in ('2026-07', '2026.7'), d.get('label'))
        check('X-Round-Result 한글 메시지 복원됨',
              isinstance(d.get('message'), str) and any('가' <= ch <= '힣' for ch in d.get('message', '')),
              d.get('message'))

    with open(dcm_path, 'rb') as f:
        r = c.post('/rounds/add',
                   data={'file': (f, 'DCM_측정.xlsx'), 'label': '2026.07'},
                   content_type='multipart/form-data')
    check('/rounds/add DCM 200', r.status_code == 200, r.data[:300])
    assert_latin1_headers(r, '/rounds/add(DCM)')

    print('[5] 라벨 정규화(canon_label) — 중복 회차가 생기지 않아야 함')
    import rounds  # noqa: E402
    for raw, want in (('2026-07', '2026.7'), ('2026.07', '2026.7'), ('2026-07-01', '2026.7'),
                      ('2026.7', '2026.7'), ('2026-01', '2026.1'), ('2025.1', '2025.1')):
        check('canon_label(%s)=%s' % (raw, want), rounds.canon_label(raw) == want, rounds.canon_label(raw))
    labels = [x['label'] if isinstance(x, dict) else x for x in rounds.list_rounds()]
    check('업로드 2건이 단일 라벨 2026.7로 병합', labels.count('2026.7') == 1 and len(labels) == 1, labels)

    print('[6] /rounds/data payload')
    r = c.get('/rounds/data')
    check('/rounds/data 200', r.status_code == 200, r.status_code)
    assert_latin1_headers(r, '/rounds/data')
    p = r.get_json(silent=True) or {}
    check('is_admin 주입됨', p.get('is_admin') is True, p.get('is_admin'))
    for k in ('surveys', 'qc_bias', 'ps_eval', 'dcm_eval', 'uploaded', 'round_labels'):
        check('payload.%s 존재' % k, k in p, sorted(p)[:12])
    check('시드 회차(2023.1) 포함', '2023.1' in (p.get('surveys') or []), (p.get('surveys') or [])[:4])
    check('업로드 회차(2026.7) 포함', '2026.7' in (p.get('round_labels') or []), p.get('round_labels'))

    print('[6b] /rounds/stats 누적 통계')
    r = c.get('/rounds/stats')
    check('/rounds/stats 200', r.status_code == 200, r.status_code)
    assert_latin1_headers(r, '/rounds/stats')
    S = r.get_json(silent=True) or {}
    for k in ('backend', 'round_labels', 'n_rounds', 'bias_summary', 'precision',
              'drift', 'drop_pattern', 'limits', 'note'):
        check('stats.%s 존재' % k, k in S, sorted(S)[:12])
    check('stats.n_rounds=1', S.get('n_rounds') == 1, S.get('n_rounds'))
    bs = (S.get('bias_summary') or {}).get('2026.7') or {}
    check('bias_summary에 uc·dcm 둘 다', set(bs) == {'uc', 'dcm'}, sorted(bs))
    uc_bs = bs.get('uc') or {}
    check('UC bias_summary 분석물질 4종', set(uc_bs) == {'TC', 'BF', 'HDL', 'LDL'}, sorted(uc_bs))
    hdl = uc_bs.get('HDL') or {}
    check('HDL 한계는 mg/dL 기준 ±1', hdl.get('limit') == 1.0 and hdl.get('unit') == 'mg/dL', hdl)
    check('TC 한계는 % 기준 ±1', (uc_bs.get('TC') or {}).get('unit') == '%', uc_bs.get('TC'))
    check('margin = |mean|/limit 일관',
          abs((hdl.get('margin') or 0) - abs(hdl.get('mean') or 0) / 1.0) < 0.002, hdl)
    dr = S.get('drift') or {}
    check('drift는 UC/DCM 분리 저장', set(dr) <= {'uc', 'dcm'} and set(dr), sorted(dr))
    check('회차 1개 → 기울기 판단 보류(None)',
          all(v.get('slope_per_round') is None for m in dr.values() for v in m.values()), dr)
    pr = (S.get('precision') or {}).get('2026.7') or {}
    check('precision에 CV·Day차 계산됨',
          (pr.get('uc') or {}).get('n_cv', 0) > 0 and (pr.get('uc') or {}).get('n_pairs', 0) > 0, pr)
    dp = S.get('drop_pattern') or {}
    check('drop_pattern 합계 = R1+R2+R3',
          dp.get('total') == sum((dp.get('counts') or {}).values()), dp)
    check('통계 note에 방법론 경고 포함', '판정' in (S.get('note') or ''), S.get('note'))

    print('[6d] 누적 추이 컷오프 — ★ 모드별(UC 전 회차 / DCM 2025.7~)')
    import rounds as R0  # noqa: E402
    check('rounds.CUM_START_DCM = 2025.7', R0.CUM_START_DCM == '2025.7', R0.CUM_START_DCM)
    check('rounds.CUM_START_UC = None(컷오프 없음)', R0.CUM_START_UC is None, R0.CUM_START_UC)
    check('cum_start_for 모드 분기',
          R0.cum_start_for('uc') is None and R0.cum_start_for('dcm') == '2025.7',
          (R0.cum_start_for('uc'), R0.cum_start_for('dcm')))
    # DCM 경계: 2025.1 제외 / 2025.7 포함. 2024.12는 '2024 하반기'라 제외.
    for lab, exp in (('2023.1', False), ('2024.7', False), ('2024.12', False), ('2025.1', False),
                     ('2025.7', True), ('2025-07-01', True), ('2026.1', True), ('2026.7', True)):
        check('DCM in_cumulative(%s)=%s' % (lab, exp),
              R0.in_cumulative(lab, mode='dcm') is exp, lab)
    # ★ UC는 컷오프가 없다 — 과거 회차도 전부 포함되어야 한다(v19 과적용 회귀 방지)
    for lab in ('2023.1', '2023.7', '2024.1', '2024.7', '2025.1', '2025.7', '2026.7'):
        check('UC in_cumulative(%s)=True' % lab, R0.in_cumulative(lab, mode='uc') is True, lab)
    cum, leg = R0.split_by_cutoff(['2023.1', '2025.1', '2025.7', '2026.7'], mode='dcm')
    check('split_by_cutoff(dcm) 분리', cum == ['2025.7', '2026.7'] and leg == ['2023.1', '2025.1'], (cum, leg))
    cum_uc, leg_uc = R0.split_by_cutoff(['2023.1', '2025.1', '2025.7'], mode='uc')
    check('split_by_cutoff(uc)는 자르지 않음', leg_uc == [] and len(cum_uc) == 3, (cum_uc, leg_uc))
    for k in ('cum_start', 'cum_start_uc', 'cum_start_dcm', 'cum_labels', 'cum_labels_uc',
              'legacy_labels', 'note_cutoff'):
        check('stats.%s 존재' % k, k in S, sorted(S)[:16])
    check('stats.cum_start 일치', S.get('cum_start') == R0.CUM_START_DCM, S.get('cum_start'))
    D = c.get('/rounds/data').get_json(silent=True) or {}
    for k in ('cum_start', 'cum_start_uc', 'cum_start_dcm', 'cum_note',
              'surveys_uc', 'surveys_cum', 'surveys_legacy',
              'ps_eval_cum', 'ps_eval_legacy', 'dcm_eval_cum', 'dcm_eval_legacy',
              'dcm_qc_cum', 'dcm_qc_legacy'):
        check('/rounds/data.%s 존재' % k, k in D, sorted(D)[:18])
    check('surveys_cum(DCM)은 컷오프 이후만',
          all(R0.in_cumulative(s, mode='dcm') for s in D.get('surveys_cum') or []), D.get('surveys_cum'))
    check('surveys_legacy(DCM)는 컷오프 이전만',
          all(not R0.in_cumulative(s, mode='dcm') for s in D.get('surveys_legacy') or []),
          D.get('surveys_legacy'))
    # ★ UC 축은 전 회차 — 2023.1이 반드시 살아 있어야 한다
    check('surveys_uc는 전 회차(2023.1 포함)',
          '2023.1' in (D.get('surveys_uc') or []) and
          set(D.get('surveys_uc') or []) == set(D.get('surveys') or []), D.get('surveys_uc'))
    check('컷오프로 자료를 삭제하지 않음(surveys 원본 유지)',
          set(D.get('surveys_cum') or []) | set(D.get('surveys_legacy') or []) == set(D.get('surveys') or []),
          (D.get('surveys'), D.get('surveys_cum'), D.get('surveys_legacy')))
    check('DCM 시드 과거 회차가 legacy로 분리됨(2023.1)',
          '2023.1' in (D.get('surveys_legacy') or []), D.get('surveys_legacy'))
    qb = (D.get('qc_bias') or {}).get('HDL') or {}
    check('qc_bias에 points_cum·points_legacy 분리',
          'points_cum' in qb and 'points_legacy' in qb, sorted(qb))
    # ★ qc_bias는 UC 계열 → legacy가 비어 있어야 한다(2023.1부터 전부 누적)
    check('qc_bias(UC)는 잘리지 않음 — legacy 비어 있음',
          not (qb.get('points_legacy') or {}) and '2023.1' in (qb.get('points_cum') or {}),
          (sorted(qb.get('points_cum') or {}), sorted(qb.get('points_legacy') or {})))
    check('qc_bias 원본 points는 그대로 보존',
          set(qb.get('points_cum') or {}) | set(qb.get('points_legacy') or {}) == set(qb.get('points') or {}),
          (sorted(qb.get('points') or {}), sorted(qb.get('points_cum') or {})))
    check('ps_eval(UC 계열)도 잘리지 않음', not (D.get('ps_eval_legacy') or {}), D.get('ps_eval_legacy'))
    check('dcm_eval(DCM)은 잘림', '2023.1' in (D.get('dcm_eval_legacy') or {}), sorted(D.get('dcm_eval_legacy') or {}))
    # drift는 컷오프 이전 DCM 회차를 추세에서 빼되, UC는 전부 살려야 한다
    def _q(an, v):
        return {'qc': [{'analyte': an, 'name': 'QC', 'day': 1, 'biasmgdl': v, 'biaspct': v}], 'samples': []}
    fake = {'2023.1': {'uc': _q('HDL', 0.3), 'dcm': _q('HDL', -0.5)},
            '2024.7': {'uc': _q('HDL', 0.2), 'dcm': _q('HDL', -0.4)},
            '2025.7': {'uc': _q('HDL', 0.1), 'dcm': _q('HDL', 0.1)},
            '2026.1': {'uc': _q('HDL', 0.0), 'dcm': _q('HDL', -1.0)}}
    import stats as ST  # noqa: E402
    dr2 = ST.drift(fake)
    check('drift(DCM)는 컷오프 이전 제외',
          (dr2.get('dcm', {}).get('HDL', {}).get('labels') or []) == ['2025.7', '2026.1'], dr2.get('dcm'))
    check('drift(UC)는 전 회차 사용',
          (dr2.get('uc', {}).get('HDL', {}).get('labels') or [])
          == ['2023.1', '2024.7', '2025.7', '2026.1'], dr2.get('uc'))
    check('drift 제외 목록은 DCM만', set(dr2.get('_excluded_legacy') or []) == {'2023.1/DCM', '2024.7/DCM'},
          dr2.get('_excluded_legacy'))
    bs2 = ST.bias_summary(fake)
    check('bias_summary: DCM 과거는 legacy 표시, 값은 보존',
          bs2['2023.1']['dcm']['HDL']['legacy'] is True
          and bs2['2023.1']['dcm']['HDL']['mean'] == -0.5, bs2['2023.1']['dcm'])
    check('bias_summary: UC 과거는 legacy 아님',
          bs2['2023.1']['uc']['HDL']['legacy'] is False, bs2['2023.1']['uc'])
    # 화면(JS)과 서버(Python)의 컷오프 판정이 어긋나면 안 된다 — 상수 동기화 확인
    _dash = open(os.path.join(ROOT, 'private', 'dashboard.html'), encoding='utf-8').read()
    check("dashboard.html의 CUM_START_DCM이 서버와 동일",
          ("const CUM_START_DCM='%s'" % R0.CUM_START_DCM) in _dash,
          [l for l in _dash.splitlines() if 'CUM_START_DCM=' in l][:1])
    check("dashboard.html의 CUM_START_UC도 컷오프 없음",
          'const CUM_START_UC=null' in _dash,
          [l for l in _dash.splitlines() if 'CUM_START_UC' in l][:1])
    # ★ UC 그래프가 다시 잘리는 회귀 방지 — 평가 경향은 DATA.ps_surveys(전 회차)를 그대로 써야 한다
    check('②탭 평가 bias 경향이 전 회차 축 사용',
          'function evalTrendChart(an){const surv=DATA.ps_surveys;' in _dash)
    check('②탭 평가 CV 경향이 전 회차 축 사용',
          'function evalCVChart(an){const surv=DATA.ps_surveys;' in _dash)
    check('⑥탭 DCM 차트만 별도 축(survDcm) 사용',
          "mk('cum_dcm',{type:'line',data:{labels:survDcm," in _dash)
    check('⑥탭 UC 차트는 전 회차 축(surv) 사용',
          "mk('cum_tcbf',{type:'line',data:{labels:surv," in _dash
          and "mk('cum_hdlldl',{type:'line',data:{labels:surv," in _dash)
    for _id in ('cumDcmCutNote', 'dcmRelCutNote'):
        check('컷오프 안내 컨테이너 %s' % _id, _dash.count(_id) >= 2, _dash.count(_id))
    check('컷오프 배지 CSS가 #t6 스코프 밖에도 정의됨',
          '\n  .cum-badge{' in _dash, '#t6 스코프에만 있으면 ②·④탭에서 색이 안 나온다')

    print('[6e] ⑥탭은 QC 경향 전용 — 검토파일 생성 없음 (사용자 확정 2026-07-28)')
    with open(dcm_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, '경향용_DCM.xlsx'), 'label': '2027.7'},
                   content_type='multipart/form-data')
    check('/rounds/add 200', r.status_code == 200, r.status_code)
    assert_latin1_headers(r, '/rounds/add(경향 전용)')
    check('/rounds/add는 JSON 반환(엑셀 아님)',
          r.mimetype == 'application/json' and r.data[:1] == b'{', (r.mimetype, r.data[:12]))
    check('검토 엑셀(zip) 아님 — 회귀 감시', r.data[:2] != b'PK', r.data[:8])
    jb = r.get_json(silent=True) or {}
    check('review_file=False 명시', jb.get('review_file') is False, jb)
    check('저장은 정상 수행', jb.get('stored') is True, jb)
    check('본문에 자동검토 패널 안내', '자동 검토' in str(jb.get('note') or ''), jb.get('note'))
    check('첨부 다운로드 헤더 없음',
          'attachment' not in str(r.headers.get('Content-Disposition') or ''),
          r.headers.get('Content-Disposition'))
    check('X-Round-Result 헤더는 유지', bool(r.headers.get('X-Round-Result')))
    c.post('/rounds/delete', data={'label': '2027.7', 'ajax': '1'})
    # /review(왼쪽 자동검토)는 종전대로 검토 엑셀을 만들어야 한다
    with open(dcm_path, 'rb') as f:
        r = c.post('/review', data={'file': (f, '자동검토_DCM.xlsx')},
                   content_type='multipart/form-data')
    check('/review는 검토 엑셀을 계속 생성', r.status_code == 200 and r.data[:2] == b'PK',
          (r.status_code, r.data[:8]))
    check('⑥탭 버튼 문구에서 검토파일 제거',
          '회차 추가 (누적 저장)' in _dash and '회차 추가 · 검토파일 생성' not in _dash)
    check('⑥탭 화면에 QC 경향 전용 안내',
          'QC bias 경향성 분석 전용' in _dash and "왼쪽 '자동 검토' 패널" in _dash)

    print('[6f] ⑦탭 CRMLN 제출 결과 선택 기준 — 수기 기록(다음 회차 참고)')
    r = c.get('/selections')
    check('/selections 200', r.status_code == 200, r.status_code)
    assert_latin1_headers(r, '/selections')
    P = r.get_json(silent=True) or {}
    for k in ('labels', 'selections', 'computed', 'reps', 'keep_n', 'note'):
        check('selections.%s 존재' % k, k in P, sorted(P))
    check('채택 2반복 규칙 명시', P.get('keep_n') == 2, P.get('keep_n'))
    check('R1–R4 목록', P.get('reps') == ['R1', 'R2', 'R3', 'R4'], P.get('reps'))
    check('기록 전용임을 note에 명시',
          '채택 로직은 이 값을 읽지 않' in str(P.get('note') or ''), P.get('note'))
    good = {'label': '2026.7', 'mode': 'dcm',
            'samples': {'CS01': {'1': {'keep': ['R2', 'R3'], 'note': 'R1 이상치'},
                                 '2': {'keep': ['R1', 'R4'], 'note': ''}},
                        'CS02': {'1': {'keep': ['R1', 'R2'], 'note': ''}}}}
    r = c.post('/selections/save', json=good)
    check('/selections/save 200', r.status_code == 200, r.data[:200])
    assert_latin1_headers(r, '/selections/save')
    check('저장 ok=True', (r.get_json(silent=True) or {}).get('ok') is True, r.get_json(silent=True))
    P2 = c.get('/selections').get_json(silent=True) or {}
    got = ((P2.get('selections') or {}).get('2026.7') or {}).get('dcm') or {}
    check('저장 후 조회 왕복', (got.get('samples') or {}).get('CS01', {}).get('1', {}).get('keep')
          == ['R2', 'R3'], got)
    check('사유 메모도 보존', (got.get('samples') or {}).get('CS01', {}).get('1', {}).get('note')
          == 'R1 이상치', got)
    check('저장 시각·작성자 기록', bool(got.get('saved_at')), got)
    check('회차 라벨 정규화 적용(2026-07 → 2026.7)',
          (c.post('/selections/save', json=dict(good, label='2026-07')).get_json(silent=True)
           or {}).get('label') == '2026.7')
    # ★ 채택 2개 규칙 위반은 서버가 거부해야 한다 — 다음 회차 참고 자료로서 의미가 없어진다
    for bad, desc in (({'CS01': {'1': {'keep': ['R1']}}}, '채택 1개'),
                      ({'CS01': {'1': {'keep': ['R1', 'R2', 'R3']}}}, '채택 3개'),
                      ({'CS01': {'1': {'keep': ['R1', 'R1']}}}, '중복 선택'),
                      ({'CS01': {'1': {'keep': ['R9', 'R2']}}}, 'R1–R4 범위 밖'),
                      ({'CS01': {'3': {'keep': ['R1', 'R2']}}}, 'Day3')):
        rr = c.post('/selections/save', json={'label': '2026.7', 'mode': 'dcm', 'samples': bad})
        check('선택 저장 거부 — %s' % desc, rr.status_code == 400, (desc, rr.status_code))
        assert_latin1_headers(rr, '/selections/save(%s)' % desc)
    check('잘못된 모드 거부',
          c.post('/selections/save', json={'label': '2026.7', 'mode': 'zz', 'samples': {}}).status_code == 400)
    check('라벨 없으면 거부',
          c.post('/selections/save', json={'label': '', 'mode': 'dcm', 'samples': {}}).status_code == 400)
    # ★ 기록이 채택 로직에 영향을 주지 않아야 한다(§0) — 저장 전후 검토 결과가 같아야 한다
    import review_engine as _RE0  # noqa: E402
    with open(dcm_path, 'rb') as f:
        _before = _RE0.summarize_round(f.read())
    c.post('/selections/save', json={'label': '2026.7', 'mode': 'dcm',
                                     'samples': {'CS01': {'1': {'keep': ['R1', 'R2']}}}})
    with open(dcm_path, 'rb') as f:
        _after = _RE0.summarize_round(f.read())
    check('수기 기록이 채택 로직에 영향 없음(§0)',
          json.dumps(_before, sort_keys=True) == json.dumps(_after, sort_keys=True))
    check('⑦탭 패널·탭 버튼 존재',
          'id="t7"' in _dash and '⑦ 제출 결과 선택 기준' in _dash)
    check('탭 전환 배열에 t7 포함', "['t1','t2','t3','t4','t5','t6','t7']" in _dash)
    check('⑦탭 초기화 훅 연결', 'window.initSel' in _dash and "b.dataset.tab==='t7'" in _dash)
    check('⑦탭에 기록 전용 경고', '이 표는 기록 전용입니다' in _dash)

    print('[6g] ⑦탭 이상치 판정 기준(선택 옵션) — ①·④탭 목록과 동기화')
    C = P.get('criteria') or {}
    check('criteria에 uc·dcm 목록', set(C) == {'uc', 'dcm'}, sorted(C))
    check('UC 기준 10종', len(C.get('uc') or []) == 10, len(C.get('uc') or []))
    check('DCM 기준 4종', len(C.get('dcm') or []) == 4, len(C.get('dcm') or []))
    check('기본값 uc=combo / dcm=median',
          (P.get('criteria_default') or {}) == {'uc': 'combo', 'dcm': 'median'},
          P.get('criteria_default'))
    # ★ 정본은 rounds.SEL_CRITERIA — ①탭(#selCrit)·④탭(#dcmSelCrit) 인라인 option과 어긋나면 안 된다.
    #   어긋나면 ⑦탭에서 고른 기준이 실제 화면 기준과 달라져 기록이 무의미해진다.
    import re as _re  # noqa: E402
    def _screen_opts(sel_id):
        m = _re.search(r'<select[^>]*id="%s"(.*?)</select>' % sel_id, _dash, _re.S)
        return _re.findall(r'<option value="([^"]+)"', m.group(1)) if m else []
    for sel_id, mode in (('selCrit', 'uc'), ('dcmSelCrit', 'dcm')):
        want = [x['value'] for x in C.get(mode) or []]
        check('%s(%s) 화면 옵션 = 서버 목록' % (sel_id, mode), _screen_opts(sel_id) == want,
              (_screen_opts(sel_id), want))
    # 민감도(방향성) 표시 — §0 경고 대상
    _uc_sens = {x['value'] for x in (C.get('uc') or []) if x.get('sensitivity')}
    check('UC 방향성 4종이 민감도로 표시',
          _uc_sens == {'hdl_high', 'hdl_low', 'bf_high', 'bf_low'}, sorted(_uc_sens))
    check('DCM 기본(median)은 민감도 아님',
          not [x for x in (C.get('dcm') or []) if x['value'] == 'median' and x.get('sensitivity')])
    # 저장·조회 왕복 — 기준이 함께 남아야 한다
    r = c.post('/selections/save', json={'label': '2026.7', 'mode': 'dcm', 'criterion': 'minpair',
                                         'samples': {'CS01': {'1': {'keep': ['R1', 'R2']}}}})
    check('기준 지정 저장 200', r.status_code == 200, r.data[:200])
    rec = (((c.get('/selections').get_json(silent=True) or {}).get('selections') or {})
           .get('2026.7') or {}).get('dcm') or {}
    check('선택 기준 저장됨', rec.get('criterion') == 'minpair', rec)
    check('기준 표시문구 함께 저장', '최소분산쌍' in str(rec.get('criterion_label') or ''), rec)
    check('민감도 플래그 저장', rec.get('sensitivity') is True, rec)
    r = c.post('/selections/save', json={'label': '2026.7', 'mode': 'dcm',
                                         'samples': {'CS01': {'1': {'keep': ['R1', 'R2']}}}})
    rec2 = (((c.get('/selections').get_json(silent=True) or {}).get('selections') or {})
            .get('2026.7') or {}).get('dcm') or {}
    check('기준 생략 시 모드 기본값(median)', rec2.get('criterion') == 'median', rec2)
    check('기본값은 민감도 아님', rec2.get('sensitivity') is False, rec2)
    rr = c.post('/selections/save', json={'label': '2026.7', 'mode': 'dcm', 'criterion': 'bf_high',
                                          'samples': {}})
    check('타 모드 기준 거부(DCM에 bf_high 없음)', rr.status_code == 400, rr.status_code)
    assert_latin1_headers(rr, '/selections/save(잘못된 기준)')
    check('없는 기준 값 거부',
          c.post('/selections/save', json={'label': '2026.7', 'mode': 'uc',
                                           'criterion': 'zzz', 'samples': {}}).status_code == 400)
    check('⑦탭 기준 드롭다운·경고 DOM', 'selCrit2' in _dash and 'selCritWarn' in _dash)
    check('⑦탭이 서버 목록으로 드롭다운 생성', 'SEL.data&&SEL.data.criteria' in _dash)
    check('민감도 선택 시 경고 문구', '방향성(민감도) 기준입니다' in _dash)

    print('[6c] T1 과거 회차 소급 누적 — 라벨 추정·연도 제한·시드 병기')
    import rounds as R  # noqa: E402
    # (1) 라벨 자동 추정: 파일명·시트명에서만 추정하며 저장하지 않는다
    cands = R.infer_labels('2025.1_HDLC_UC_측정결과.xlsx', ['결과정리', '2025년 1월 검토'])
    labs = [x['label'] for x in cands]
    check('파일명에서 라벨 추정', '2025.1' in labs, cands)
    check('추정 근거(source)를 함께 반환', all(x.get('source') and x.get('matched') for x in cands), cands)
    check('상/하반기 표기도 추정', '2024.7' in [x['label'] for x in R.infer_labels('2024 하반기.xlsx')],
          R.infer_labels('2024 하반기.xlsx'))
    check('추정 불가 파일은 빈 목록', R.infer_labels('측정결과.xlsx', ['결과정리']) == [],
          R.infer_labels('측정결과.xlsx', ['결과정리']))
    check('13월 등 잘못된 월은 추정 안 함', R.infer_labels('2025.13_x.xlsx') == [],
          R.infer_labels('2025.13_x.xlsx'))
    check('숫자 오인 방지(2025.1234)', R.infer_labels('2025.1234.xlsx') == [],
          R.infer_labels('2025.1234.xlsx'))
    check('YYYY-MM-DD도 반기로 정규화',
          [x['label'] for x in R.infer_labels('2024-07-15.xlsx')] == ['2024.7'],
          R.infer_labels('2024-07-15.xlsx'))
    check('같은 라벨 여러 근거 → n_sources 집계',
          all(x['n_sources'] == 2 for x in R.infer_labels('2025.1.xlsx', ['2025년 1월'])),
          R.infer_labels('2025.1.xlsx', ['2025년 1월']))

    # Postgres 경로 재현 — db.rounds_load()는 레코드 최상위에 reference를 넣지 않고
    # 요약(data JSONB) 안에만 넣는다. 그 형태에서도 참고용 판정이 되어야 한다.
    import stats as _S
    pg_like = {'2024.1': {'label': '2024.1', 'by': {'uc': 'admin'}, 'uc': {
        'mode': 'uc', 'reference': True, 'n_qc': 1, 'n_exceed': 0, 'n_samples': 1,
        'qc': [{'analyte': 'TC', 'biaspct': 0.2, 'biasmgdl': None, 'name': 'NIST', 'day': 1, 'ok': True}],
        'samples': [{'name': 'CS01', 'day': 1, 'drop': 2, 'keep': [1, 3],
                     'BF': 1, 'HDL': 50, 'LDL': 1, 'cvL': 0.3}]}}}
    check('Postgres 형태(요약 안에만 reference)에서도 참고용 인식',
          R.is_reference(pg_like['2024.1']) is True)
    check('일반 회차는 참고용 아님', R.is_reference({'uc': {'mode': 'uc'}}) is False)
    _d = _S.drift(pg_like)
    check('Postgres 형태에서도 드리프트가 참고용 제외',
          _d.get('_excluded_reference') == ['2024.1/UC'], _d)
    check('참고용만 있으면 추세 계열이 비어야 함', set(_d) == {'_excluded_reference'}, sorted(_d))
    check('bias_summary 행에 reference 표시',
          _S.bias_summary(pg_like)['2024.1']['uc']['TC']['reference'] is True)

    # (2) 최근 3년 제한 — 시점 불확실 시 min_backfill_year 이전은 거부
    lo = R.min_backfill_year()
    check('최근 3년 하한 = 올해-2', lo == __import__('datetime').date.today().year - 2, lo)
    ok_old, msg_old = R.year_guard('%d.1' % (lo - 1), date_certain=False)
    check('시점 불확실 + 3년 초과 과거 → 거부', ok_old is False and str(lo) in msg_old, msg_old)
    ok_cert, _ = R.year_guard('%d.1' % (lo - 1), date_certain=True)
    check('시점 확인됨 체크 시 허용', ok_cert is True)
    check('범위 내 연도는 허용', R.year_guard('%d.7' % lo)[0] is True)
    check('미래 연도 거부', R.year_guard('2099.1')[0] is False)

    # (3) /rounds/preview — 저장하지 않는다
    before = set(R.load_store())
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/preview', data={'file': (f, '2025.1_과거측정.xlsx')},
                   content_type='multipart/form-data')
    check('/rounds/preview 200', r.status_code == 200, r.data[:200])
    assert_latin1_headers(r, '/rounds/preview')
    P = r.get_json(silent=True) or {}
    check('preview saved=False', P.get('saved') is False, P.get('saved'))
    check('preview는 저장하지 않음', set(R.load_store()) == before, sorted(set(R.load_store()) - before))
    check('preview 추정 라벨 제시', P.get('suggested_label') == '2025.1', P.get('suggested_label'))
    check('preview 기존 라벨 상태 포함', isinstance(P.get('status'), dict) and 'exists' in P['status'], P.get('status'))
    check('preview 시드 대조 포함', isinstance(P.get('seed_compare'), dict), P.get('seed_compare'))
    check('preview note에 사용자 확인 요구 명시', '확정' in (P.get('note') or ''), P.get('note'))

    # (4) 확인(confirm) 없이는 소급 저장 불가
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, 'a.xlsx'), 'label': '2025.1', 'reference': '1'},
                   content_type='multipart/form-data')
    check('confirm 없는 소급 저장 거부 400', r.status_code == 400, r.status_code)
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, 'a.xlsx'), 'label': '%d.1' % (lo - 1),
                                        'reference': '1', 'confirm': '1'},
                   content_type='multipart/form-data')
    check('3년 초과 과거 소급 거부 400', r.status_code == 400, r.data[:200])

    # (5) 정상 소급 저장 — 참고용 표시 + 검토 엑셀 미생성(JSON 응답)
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, 'a.xlsx'), 'label': '%d.1' % lo,
                                        'reference': '1', 'confirm': '1'},
                   content_type='multipart/form-data')
    check('소급 저장 200', r.status_code == 200, r.data[:200])
    assert_latin1_headers(r, '/rounds/add(소급)')
    J = r.get_json(silent=True) or {}
    check('소급 응답은 JSON(검토 엑셀 미생성)', J.get('stored') is True and J.get('reference') is True, J)
    st = R.load_store().get('%d.1' % lo) or {}
    check('저장 레코드에 reference 표시', R.is_reference(st) is True, st.get('reference'))

    # (6) 시드 값 보존 — 덮어쓰지 않고 병기
    P2 = R.dashboard_payload()
    seedpts = (P2['qc_bias']['NIST'] or {}).get('points_seed') or {}
    seed_raw = (R.load_seed()['qc_bias']['NIST'] or {}).get('points') or {}
    check('시드 points_seed가 원본과 동일(덮어쓰기 없음)',
          all(abs(seedpts[k] - seed_raw[k]) < 1e-9 for k in seed_raw), 'seed 변형됨')
    check('업로드 값은 points_upload에 별도 보관',
          isinstance(P2['qc_bias']['NIST'].get('points_upload'), dict), P2['qc_bias']['NIST'].keys())
    check('conflicts 목록 존재', isinstance(P2.get('conflicts'), list), type(P2.get('conflicts')))
    check('payload에 reference_labels', '%d.1' % lo in (P2.get('reference_labels') or []),
          P2.get('reference_labels'))
    check('payload note_seed에 병기 원칙 명시', '덮어쓰지' in (P2.get('note_seed') or ''), P2.get('note_seed'))
    for cf in (P2.get('conflicts') or []):
        check('conflict 항목에 시드·업로드 값 모두 포함',
              cf.get('seed') is not None and cf.get('upload') is not None, cf)
        break

    # (7) 참고용 소급은 드리프트(추세) 계산에서 제외
    S2 = json.loads(c.get('/rounds/stats').data.decode('utf-8'))
    check('stats에 reference_labels 노출', '%d.1' % lo in (S2.get('reference_labels') or []),
          S2.get('reference_labels'))
    ex = (S2.get('drift') or {}).get('_excluded_reference') or []
    check('드리프트에서 참고용 소급 제외', any(('%d.1' % lo) in e for e in ex), ex)
    check('stats note_reference에 참고용 경고', '참고용' in (S2.get('note_reference') or ''),
          S2.get('note_reference'))
    c.post('/rounds/delete', data={'label': '%d.1' % lo, 'ajax': '1'})

    print('[6d] HDL-C DCM — Day2 열 자동 탐지 · n중 2개 채택')
    import review_engine as RE  # noqa: E402
    # (1) median 일반화 — n=3은 종전(sorted[1])과 동일해야 하고, n=4는 가운데 두 값 평균
    check('median n=3 회귀', RE._median([3, 1, 2]) == 2 and RE._median([5, 5, 9]) == 5)
    check('median n=4 = 가운데 2개 평균', abs(RE._median([10, 10.1, 10.2, 10.3]) - 10.15) < 1e-9,
          RE._median([10, 10.1, 10.2, 10.3]))
    # (2) 채택 규칙
    d, k, s, _cv = RE._dcm_pick([1, 2, 3])
    check('n=3 → 1개 제외·2개 채택', d == [0] and k == [1, 2], (d, k))
    d, k, s, _cv = RE._dcm_pick([10, 10.1, 10.2, 10.3])
    check('n=4 → 극단 2개 제외(순위 중앙 2개 채택)', d == [0, 3] and k == [1, 2], (d, k))
    # 중복값이 있어도 동일값 2개를 채택해 CV를 0으로 만들지 않아야 한다(§0)
    d, k, s, cv = RE._dcm_pick([50.935, 50.631, 50.631, 50.813])
    check('중복값 — 인위적 CV 0 회피', cv > 0.1 and abs(s - 50.722) < 1e-6, (k, s, cv))
    d, k, s, cv = RE._dcm_pick([48.212, 48.031, 47.911, 47.911])
    check('중복값 — 채택값이 순위 중앙 평균', abs(s - 47.971) < 1e-6, (k, s))
    # (3) 채택은 정밀도 기준뿐 — 값의 크기 순으로 고르지 않는다
    d, k, s, _cv = RE._dcm_pick([100.0, 50.0, 50.1, 50.2])
    check('이상치 1개는 반드시 제외', 0 in d, (d, k))
    # (4) Day 블록 열 자동 탐지 — Day2가 한 칸 밀린 레이아웃에서도 잡아야 한다
    import openpyxl as _ox  # noqa: E402
    wb = _ox.Workbook(); sh = wb.active
    for _c, _t in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3'), (8, 'R4'),
                   (17, 'A.value'), (18, 'R1'), (19, 'R2'), (20, 'R3'), (21, 'R4')]:
        sh.cell(RE.DCM_HEADER_ROW, _c, _t)
    cols = RE._dcm_day_cols(sh)
    check('Day1 열 탐지', cols[1]['a'] == 4 and cols[1]['r'] == [5, 6, 7, 8], cols[1])
    check('Day2 열 탐지(밀린 레이아웃)', cols[2]['a'] == 17 and cols[2]['r'] == [18, 19, 20, 21], cols[2])
    wb2 = _ox.Workbook(); sh2 = wb2.active
    for _c, _t in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3'),
                   (16, 'A.value'), (17, 'R1'), (18, 'R2'), (19, 'R3')]:
        sh2.cell(RE.DCM_HEADER_ROW, _c, _t)
    cols2 = RE._dcm_day_cols(sh2)
    check('구 레이아웃(3반복)도 그대로 탐지', cols2[2]['a'] == 16 and cols2[2]['r'] == [17, 18, 19], cols2[2])
    check('헤더 없으면 종전 하드코딩으로 폴백',
          RE._dcm_day_cols(_ox.Workbook().active)[2]['a'] == 16)
    # (4b) T5 — 세로형(Day2 stacked) 배열 탐지. 가로형은 row0가 같고, 세로형은 row0가 달라야 한다.
    check('가로형은 Day1·Day2 row0가 같다', cols[1]['row0'] == cols[2]['row0'] == 5,
          (cols[1]['row0'], cols[2]['row0']))
    for _hdr2 in (16, 17, 20):     # Day2 헤더 간격은 회차마다 다르다 — 고정 간격 가정 금지
        wb3 = _ox.Workbook(); sh3 = wb3.active
        for _r in (RE.DCM_HEADER_ROW, _hdr2):
            for _c, _t in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3'), (8, 'R4')]:
                sh3.cell(_r, _c, _t)
        cols3 = RE._dcm_day_blocks(sh3)
        check('세로형 Day2 탐지(헤더 %d행)' % _hdr2,
              cols3[2]['row0'] == _hdr2 + 1 and cols3[2]['a'] == 4 and cols3[1]['row0'] == 5,
              {k: (v['hdr'], v['row0'], v['a']) for k, v in cols3.items()})
    # Day1만 있는 측정지에서 없는 Day2를 하드코딩으로 지어내면 안 된다(§0)
    wb4 = _ox.Workbook(); sh4 = wb4.active
    for _c, _t in [(4, 'A.value'), (5, 'R1'), (6, 'R2'), (7, 'R3')]:
        sh4.cell(RE.DCM_HEADER_ROW, _c, _t)
    check('Day1만 있으면 Day2를 추측하지 않음', 2 not in RE._dcm_day_blocks(sh4),
          list(RE._dcm_day_blocks(sh4)))
    # (5) 요약 결과 구조 — drop이 리스트이고 Day1·Day2가 모두 들어와야 한다
    with open(dcm_path, 'rb') as f:
        sm = RE.summarize_round(f.read())
    check('DCM 자동 감지', sm.get('mode') == 'dcm', sm.get('mode'))
    check('Day1·Day2 모두 파싱(조용한 누락 없음)',
          {x['day'] for x in sm['samples']} == {1, 2}, sorted({x['day'] for x in sm['samples']}))
    check('QC도 Day1·Day2 모두', {q['day'] for q in sm['qc']} == {1, 2},
          sorted({q['day'] for q in sm['qc']}))
    check('drop은 리스트', all(isinstance(x['drop'], list) for x in sm['samples']))
    check('채택은 항상 2개', all(len(x['keep']) == 2 for x in sm['samples']))
    check('제외 개수 = n_reps - 2',
          all(len(x['drop']) == x['n_reps'] - 2 for x in sm['samples']),
          [(x['n_reps'], x['drop']) for x in sm['samples']])
    check('drop·keep이 서로 겹치지 않음',
          all(not (set(x['drop']) & set(x['keep'])) for x in sm['samples']))
    # (5b) 4반복·Day2 밀림 레이아웃 fixture를 업로드 경로로 end-to-end 검증
    dcm4_path = make_fixture.build_dcm4(os.path.join(fx, 'fixture_DCM4.xlsx'))
    with open(dcm4_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, '4반복_DCM.xlsx'), 'label': '2027.1'},
                   content_type='multipart/form-data')
    check('4반복 DCM 업로드 200', r.status_code == 200, r.data[:200])
    assert_latin1_headers(r, '/rounds/add(4반복 DCM)')
    with open(dcm4_path, 'rb') as f:
        sm4 = RE.summarize_round(f.read())
    check('4반복 fixture: Day1·Day2 모두 파싱',
          {x['day'] for x in sm4['samples']} == {1, 2}, sorted({x['day'] for x in sm4['samples']}))
    check('4반복 fixture: n_reps=4', all(x['n_reps'] == 4 for x in sm4['samples']),
          [x['n_reps'] for x in sm4['samples']])
    check('4반복 fixture: 2개씩 제외', all(len(x['drop']) == 2 for x in sm4['samples']))
    check('4반복 fixture: Day2 QC 이름이 Day2 열에서 읽힘',
          any(q['name'] == 'NIST2' and q['day'] == 2 for q in sm4['qc']),
          [(q['name'], q['day']) for q in sm4['qc']])
    # CS04는 중복값이 순위 중앙이 아닌 위치에 있다 → 중복 2개를 채택해 CV가 0이 되면 안 된다.
    cs04 = [x for x in sm4['samples'] if x['name'] == 'CS04']
    check('중복값이 순위 중앙이 아니면 CV가 0으로 붕괴하지 않음',
          len(cs04) == 2 and all(x['cv'] > 0 for x in cs04),
          [(x['day'], x['reps'], x['keep'], x['cv']) for x in cs04])
    c.post('/rounds/delete', data={'label': '2027.1', 'ajax': '1'})

    # (5b-2) T5 — 세로형(Day2 stacked) 측정지를 업로드 경로로 end-to-end 검증.
    #   2025.7(`Sheet2`, Day2 헤더 17행)·2026.1(`결과 취합`, 16행) 실제 배열을 재현한 fixture.
    #   ★ Day1/Day2 값이 서로 다르므로, 행 하드코딩으로 돌아가면 Day2가 Day1 사본이 되어 여기서 걸린다.
    for _sheet, _h2, _tag in (('결과 취합', 16, '결과취합'), ('Sheet2', 17, 'Sheet2')):
        sv = make_fixture.build_dcm_stacked(
            os.path.join(fx, 'fixture_DCM_세로_%s.xlsx' % _tag), _sheet, _h2)
        with open(sv, 'rb') as f:
            svb = f.read()
        svwb = _ox.load_workbook(io.BytesIO(svb))
        check('세로형[%s]: 측정 시트 자동 인식' % _sheet,
              RE.find_ms_sheet(svwb) == _sheet, RE.find_ms_sheet(svwb))
        check('세로형[%s]: DCM으로 판별' % _sheet, RE._is_dcm(svwb[_sheet]))
        smv = RE.summarize_round(svb)
        check('세로형[%s]: Day1·Day2 모두 파싱' % _sheet,
              {x['day'] for x in smv['samples']} == {1, 2},
              sorted({x['day'] for x in smv['samples']}))
        check('세로형[%s]: QC도 Day1·Day2 모두' % _sheet,
              {q['day'] for q in smv['qc']} == {1, 2}, sorted({q['day'] for q in smv['qc']}))
        check('세로형[%s]: 검체 4개' % _sheet, smv['n_samples'] == 4, smv['n_samples'])
        # Day1·Day2 채택값이 전부 같다면 세로 탐색이 실패해 같은 행을 두 번 읽은 것이다.
        _d1 = {x['name']: x['HDL'] for x in smv['samples'] if x['day'] == 1}
        _d2 = {x['name']: x['HDL'] for x in smv['samples'] if x['day'] == 2}
        check('세로형[%s]: Day2가 Day1 사본이 아님' % _sheet,
              _d1 and _d2 and any(_d1[k] != _d2.get(k) for k in _d1), (_d1, _d2))
        with open(sv, 'rb') as f:
            r = c.post('/rounds/add', data={'file': (f, '세로형_DCM.xlsx'), 'label': '2027.7'},
                       content_type='multipart/form-data')
        check('세로형[%s]: 업로드 200' % _sheet, r.status_code == 200, r.data[:200])
        assert_latin1_headers(r, '/rounds/add(세로형 DCM %s)' % _sheet)
        c.post('/rounds/delete', data={'label': '2027.7', 'ajax': '1'})
        # 검토파일 생성까지 — 선택 시트가 세로형 원본 행을 참조해야 한다
        svout, _svm = RE.process(svb, label='2027.7')
        svwb2 = _ox.load_workbook(io.BytesIO(svout))
        _svsel = RE.sel_sheet_name('2027.7')
        check('세로형[%s]: 검토파일에 결과선택 시트' % _sheet,
              _svsel in svwb2.sheetnames, svwb2.sheetnames)
        _sel = svwb2[_svsel]
        # Day2 표(V열~)의 첫 행 수식이 Day2 실제 시작 행(_h2+1)을 가리켜야 한다
        _f = str(_sel.cell(15, 2 + 20 + 3).value or '')
        check('세로형[%s]: Day2 수식이 %d행을 참조' % (_sheet, _h2 + 1),
              ('%d' % (_h2 + 1)) in _f and _sheet in _f, _f)

    # (5c) DCM 검토파일도 UC와 동일하게 [검토_가이드]·[결과선택] 시트를 포함해야 한다
    import openpyxl as _ox2  # noqa: E402
    with open(dcm_path, 'rb') as f:
        dcm_out, _dmeta = RE.process(f.read(), label='2026.7')
    dwb = _ox2.load_workbook(io.BytesIO(dcm_out))
    _dsel = RE.sel_sheet_name('2026.7')
    check('DCM 검토파일에 검토_가이드 시트', RE.DCM_GUIDE in dwb.sheetnames, dwb.sheetnames)
    check('DCM 검토파일에 결과선택 시트', _dsel in dwb.sheetnames, dwb.sheetnames)
    check('DCM 검토파일에 검토 시트', RE.DCM_SHEET in dwb.sheetnames, dwb.sheetnames)
    sel = dwb[_dsel]
    check('선택 시트 C4에 옵션 드롭다운',
          any(str(d.sqref) == 'C4' and d.type == 'list' for d in sel.data_validations.dataValidation),
          [(str(d.sqref), d.type) for d in sel.data_validations.dataValidation])
    check('기본 옵션이 정밀도 기준', sel['C4'].value == RE.DCM_DEFAULT_OPTION, sel['C4'].value)
    check('선택 시트가 측정 시트를 수식으로 참조',
          isinstance(sel['E15'].value, str) and sel['E15'].value.startswith('=IF('), sel['E15'].value)
    check('조건부 서식(채택/제외/검증) 등록', len(list(sel.conditional_formatting)) >= 4,
          len(list(sel.conditional_formatting)))
    # 헬퍼 열이 겹치면 Day1의 '기본 drop2'가 Day2의 n 수식에 덮여 조용히 오답이 된다(실제 발생)
    _h1, _h2, _hw, _oc = RE.H_DCM_SEL
    check('Day1·Day2 헬퍼 열이 겹치지 않음', _h1 + _hw <= _h2, RE.H_DCM_SEL)
    check('헬퍼가 옵션 목록 열을 침범하지 않음', _h2 + _hw <= _oc, RE.H_DCM_SEL)
    # 서버 계산값이 검증용으로 기록되어 있어야 한다
    srv_cells = [sel.cell(15 + (sr - 5), 2 + 16).value for sr in (9, 10, 11, 12)]
    check('서버 계산값이 검증 열에 기록됨', all(isinstance(v, (int, float)) for v in srv_cells), srv_cells)
    # ★ 조건부 서식 색상 회귀 방지 — dxf는 fgColor가 아니라 bgColor로 렌더된다.
    #   fgColor만 넣으면 Excel에서 채택 셀 노란색이 아예 보이지 않는다(실제 발생).
    #   6자리 RGB를 주면 알파가 00(투명)으로 저장되므로 8자리 ARGB여야 한다.
    import zipfile as _zip, re as _re  # noqa: E402
    _sx = _zip.ZipFile(io.BytesIO(dcm_out)).read('xl/styles.xml').decode('utf-8')
    _dx = _re.search(r'<dxfs.*?</dxfs>', _sx, _re.S)
    _fills = _re.findall(r'<patternFill[^>]*>(.*?)</patternFill>', _dx.group(0) if _dx else '', _re.S)
    check('조건부 서식 채우기가 하나 이상', len(_fills) >= 3, len(_fills))
    check('dxf 채우기가 bgColor를 사용(fgColor 단독 금지)',
          all('bgColor' in f for f in _fills), _fills[:2])
    check('dxf 색상이 8자리 ARGB(알파 FF)',
          all(_re.search(r'bgColor rgb="FF[0-9A-Fa-f]{6}"', f) for f in _fills), _fills[:2])
    check('채택 셀 노란색(FFF2A8)이 등록됨',
          any('FFFFF2A8' in f for f in _fills), _fills)
    check('제외 셀 회색(F2F2F2)이 등록됨', any('FFF2F2F2' in f for f in _fills), _fills)
    check('제외 셀에 취소선', 'strike' in (_dx.group(0) if _dx else ''))

    gd = dwb[RE.DCM_GUIDE]
    gtxt = ' '.join(str(c.value) for r in gd.iter_rows() for c in r if c.value)
    check('가이드에 판정 기준(±1 mg/dL) 명시', '±1 mg/dL' in gtxt)
    check('가이드에 동점 처리 설명', '동점' in gtxt and '중앙' in gtxt)
    check('가이드에 유리한 선택 지양 원칙', '유리하게 만들기 위한 선택은 지양' in gtxt)
    check('가이드에 검증 열 안내', '불일치' in gtxt)
    check('가이드가 회차 시트명을 그대로 안내', _dsel in gtxt, gtxt[:120])

    # ------------------------------------------------------------------
    # (5c-2) T7 — 선택 시트명이 회차 라벨을 따라간다
    #   종전에는 '2026.7_결과선택'으로 상수 고정이라 2026.1 측정지를 올려도 시트명·제목이
    #   '2026-07 회차'로 나왔다. 표시 문제지만 과거 회차 검토파일에서 오해 소지가 컸다.
    #   ★ 라벨은 표시 전용이며 채택값·계산에는 관여하지 않는다(§0) — 아래에서 함께 검증한다.
    # ------------------------------------------------------------------
    check('sel_sheet_name: 라벨 있으면 접두', RE.sel_sheet_name('2026.1') == '2026.1_결과선택',
          RE.sel_sheet_name('2026.1'))
    check('sel_sheet_name: 라벨 없으면 회차 표기 없음', RE.sel_sheet_name('') == '결과선택',
          RE.sel_sheet_name(''))
    check('canon_round_label 정규화', RE.canon_round_label('2026-07-01') == '2026.7'
          and RE.canon_round_label('2026-01') == '2026.1',
          (RE.canon_round_label('2026-07-01'), RE.canon_round_label('2026-01')))
    check('infer_round_label: 파일명에서 추정',
          RE.infer_round_label('CRMLN_2026-01_DCM측정.xlsx') == '2026.1',
          RE.infer_round_label('CRMLN_2026-01_DCM측정.xlsx'))
    check('infer_round_label: 시트명에서 추정',
          RE.infer_round_label('무제.xlsx', ['2025년 7월 결과']) == '2025.7',
          RE.infer_round_label('무제.xlsx', ['2025년 7월 결과']))
    check('infer_round_label: 근거 없으면 빈 문자열',
          RE.infer_round_label('무제.xlsx', ['Sheet2']) == '',
          repr(RE.infer_round_label('무제.xlsx', ['Sheet2'])))

    with open(dcm_path, 'rb') as f:
        _dcm_raw = f.read()
    _o_lab, _ = RE.process(_dcm_raw, label='2026.1')          # 사용자 확정 라벨
    _o_inf, _ = RE.process(_dcm_raw, filename='2025.7_DCM.xlsx')  # 파일명 추정
    _o_non, _ = RE.process(_dcm_raw)                           # 근거 없음
    _w_lab = _ox2.load_workbook(io.BytesIO(_o_lab))
    _w_inf = _ox2.load_workbook(io.BytesIO(_o_inf))
    _w_non = _ox2.load_workbook(io.BytesIO(_o_non))
    check('T7: 확정 라벨이 시트명에 반영', '2026.1_결과선택' in _w_lab.sheetnames, _w_lab.sheetnames)
    check('T7: 옛 고정명(2026.7_결과선택)이 남지 않음',
          '2026.7_결과선택' not in _w_lab.sheetnames, _w_lab.sheetnames)
    check('T7: 파일명 추정이 시트명에 반영', '2025.7_결과선택' in _w_inf.sheetnames, _w_inf.sheetnames)
    check('T7: 근거 없으면 회차 없는 시트명', '결과선택' in _w_non.sheetnames, _w_non.sheetnames)
    _o_both, _ = RE.process(_dcm_raw, filename='2025.7_x.xlsx', label='2026.1')
    _w_both = _ox2.load_workbook(io.BytesIO(_o_both))
    check('T7: 확정 라벨이 파일명 추정보다 우선',
          '2026.1_결과선택' in _w_both.sheetnames and '2025.7_결과선택' not in _w_both.sheetnames,
          _w_both.sheetnames)
    _ttl = str(_w_lab[RE.DCM_SHEET].cell(1, 1).value or '')
    check('T7: 검토 시트 제목이 회차를 따라감', '2026년 1월' in _ttl, _ttl)
    check('T7: 제목에 옛 회차(2026년 7월)가 남지 않음', '2026년 7월' not in _ttl, _ttl)
    _ttl_non = str(_w_non[RE.DCM_SHEET].cell(1, 1).value or '')
    check('T7: 회차를 모르면 제목에 회차를 지어내지 않음',
          '년' not in _ttl_non.split('HDL')[0], _ttl_non)
    # ★ 라벨은 표시 전용 — 채택값·요약이 라벨에 따라 달라지면 안 된다
    check('T7: 라벨이 달라도 요약(계산값) 동일',
          RE.summarize_round(_dcm_raw) == RE.summarize_round(_dcm_raw), 'summarize_round 결정성')
    _sv_lab = [_w_lab['2026.1_결과선택'].cell(15 + i, 2 + 16).value for i in range(8)]
    _sv_non = [_w_non['결과선택'].cell(15 + i, 2 + 16).value for i in range(8)]
    check('T7: 시트명이 달라도 서버 계산값 동일', _sv_lab == _sv_non, (_sv_lab, _sv_non))

    # UC 경로도 같은 규칙을 따라야 한다(템플릿 시트를 rename하는 별도 경로라 따로 검사)
    with open(uc_path, 'rb') as f:
        _uc_raw = f.read()
    _u_out, _ = RE.process(_uc_raw, filename='CRMLN 2025년 1월 UC 측정.xlsx')
    _u_wb = _ox2.load_workbook(io.BytesIO(_u_out))
    check('T7(UC): 파일명 추정이 시트명에 반영', '2025.1_결과선택' in _u_wb.sheetnames, _u_wb.sheetnames)
    check('T7(UC): 템플릿 원본 선택 시트명이 남지 않음',
          RE.SEL_SRC not in _u_wb.sheetnames, _u_wb.sheetnames)
    _u_ttl = str(_u_wb[RE.UC_SHEET].cell(1, 1).value or '')
    check('T7(UC): 검토 시트 제목이 회차를 따라감', '2025년 1월' in _u_ttl, _u_ttl)
    _ug = ' '.join(str(c.value) for r in _u_wb['검토_가이드'].iter_rows() for c in r if c.value)
    check('T7(UC): 가이드 본문의 시트명도 갱신',
          '2025.1_결과선택' in _ug and '2026.7_결과선택' not in _ug, _ug[:150])
    _u_none, _ = RE.process(_uc_raw)
    _u_wb2 = _ox2.load_workbook(io.BytesIO(_u_none))
    check('T7(UC): 근거 없으면 회차 없는 시트명', '결과선택' in _u_wb2.sheetnames, _u_wb2.sheetnames)

    # ★ 사용자 시트 보호 — DCM 경로는 업로드 워크북을 그대로 편집하므로, 이름만 보고 지우면
    #   사용자가 만든 '*_결과선택' 시트가 사라진다. 서명이 없는 시트는 남아 있어야 한다.
    _uwb = _ox2.load_workbook(io.BytesIO(_dcm_raw))
    _mine = _uwb.create_sheet('2024.1_결과선택')
    _mine['A1'] = '검토자 개인 메모 — 앱이 만든 시트 아님'
    _bio = io.BytesIO(); _uwb.save(_bio)
    _o_keep, _ = RE.process(_bio.getvalue(), label='2026.1')
    _w_keep = _ox2.load_workbook(io.BytesIO(_o_keep))
    check('T7: 사용자가 만든 결과선택 시트는 보존', '2024.1_결과선택' in _w_keep.sheetnames, _w_keep.sheetnames)
    check('T7: 사용자 시트 내용도 그대로',
          _w_keep['2024.1_결과선택']['A1'].value == '검토자 개인 메모 — 앱이 만든 시트 아님',
          _w_keep['2024.1_결과선택']['A1'].value)
    # 반대로 앱이 만든 옛 회차 선택 시트는 재생성 시 정리되어야 한다(중복 방지)
    _o_re, _ = RE.process(_o_lab, label='2025.7')   # 이미 '2026.1_결과선택'이 든 검토파일을 재처리
    _w_re = _ox2.load_workbook(io.BytesIO(_o_re))
    check('T7: 앱이 만든 옛 회차 선택 시트는 정리',
          '2025.7_결과선택' in _w_re.sheetnames and '2026.1_결과선택' not in _w_re.sheetnames,
          _w_re.sheetnames)

    # (5d) DCM 내부 QC bias 시드(2025.7·2026.1·2026.7) — 평가 bias와 병기되어야 한다
    P3 = R.dashboard_payload()
    seedqc = P3.get('dcm_qc_seed') or {}
    check('DCM QC 시드에 2025.7·2026.1·2026.7', {'2025.7', '2026.1', '2026.7'} <= set(seedqc), sorted(seedqc))
    check('2026.1 내부 QC가 음(−) 방향', seedqc.get('2026.1', 0) < 0, seedqc.get('2026.1'))
    check('2026.1 평가도 음(−) 방향(대응 확인)',
          (P3['dcm_eval'].get('2026.1') or {}).get('bias', 0) < 0)
    paired = [s for s in P3['surveys'] if s in (P3.get('dcm_qc_upload') or {}) and s in P3['dcm_eval']]
    check('내부 QC·평가 짝 데이터 2개 이상', len(paired) >= 2, paired)
    check('시드 QC는 별도 보존(업로드가 덮어쓰지 않음)', isinstance(P3.get('dcm_qc_seed'), dict))
    check('업로드 전용 계열도 분리 제공', isinstance(P3.get('dcm_qc_upload_only'), dict))

    # (6) drop 리스트를 stats가 집계할 수 있어야 한다(구 정수 형식도 호환)
    import stats as _S2  # noqa: E402
    mixed = {'2026.7': {'label': '2026.7',
                        'uc': {'mode': 'uc', 'qc': [], 'samples': [{'drop': 2, 'n_reps': 3}]},
                        'dcm': {'mode': 'dcm', 'qc': [], 'samples': [{'drop': [1, 4], 'n_reps': 4}]}}}
    dp = _S2.drop_pattern(mixed)
    check('drop 정수·리스트 혼재 집계', dp['total'] == 3, dp)
    check('R4까지 집계', dp['counts'].get('R4') == 1, dp['counts'])
    check('균등 기대치는 실제 반복 개수 기준', dp['n_reps'] == 4 and abs(dp['expected_each'] - 0.75) < 1e-9, dp)

    print('[7] /rounds/delete (관리자, ajax)')
    r = c.post('/rounds/delete', data={'label': '2026.7', 'mode': 'dcm', 'ajax': '1'})
    check('/rounds/delete ajax 200', r.status_code == 200, r.data[:200])
    check('/rounds/delete ok=True', (r.get_json(silent=True) or {}).get('ok') is True, r.data[:200])
    assert_latin1_headers(r, '/rounds/delete')

    print('[8] 잘못된 입력 방어')
    r = c.post('/rounds/add', data={'file': (open(uc_path, 'rb'), 'x.txt')},
               content_type='multipart/form-data')
    check('비-xlsx 거부 400', r.status_code == 400, r.status_code)
    with open(uc_path, 'rb') as f:
        r = c.post('/rounds/add', data={'file': (f, 'a.xlsx'), 'label': ''},
                   content_type='multipart/form-data')
    check('라벨 누락 거부 400', r.status_code == 400, r.status_code)

    shutil.rmtree(tmp, ignore_errors=True)

    print('\n=== %d개 검사 중 실패 %d ===' % (CHECKS[0], len(FAILS)))
    if FAILS:
        for f in FAILS:
            print('  - %s' % f)
        return 1
    print('전부 통과')
    return 0


if __name__ == '__main__':
    sys.exit(main())
